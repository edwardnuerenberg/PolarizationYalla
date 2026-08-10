# --- Standard libs
import os, sys, glob, re, json, math
import shutil, subprocess, tempfile, zipfile
from datetime import datetime

# --- Tk GUI (bring this in early)
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, colorchooser

# --- Data
import pandas as pd
import numpy as np

# --- Matplotlib: force an interactive backend BEFORE importing pyplot
import matplotlib
matplotlib.use("TkAgg", force=True)          # must come before pyplot
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle  # <- correct import for the white panel
matplotlib.rcParams.update({
    "axes.titlesize": 20,
    "axes.labelsize": 16,
    "xtick.labelsize": 16,
    "ytick.labelsize": 16,
    "legend.fontsize": 14,
})

# ----------------------------
# Globals
# ----------------------------
data_dict = {}        # {folder_name: DataFrame}
current_at_06V = {}   # {folder_name: {... results per Pol ...}}

# Per-folder verdict on the raw-file layout, carried from scrape_data() all the way into
# the results table and the Excel/CSV export so a fallback can never pass unnoticed.
# {folder_name: "OK" | "OK — …" | "Check data! (…)" | "Missing"}
pol_data_structure = {}
# {folder_name: {"used": run number actually read, "unused": [other run numbers present]}}
pol_data_runs = {}
# Why a folder produced no metrics, filled by extract_current(). Kept apart from the file
# verdict above so re-running the extraction recomputes it instead of overwriting what
# scrape_data() found out about the files.
pol_data_notes = {}
DATA_STRUCTURE_COLUMN = "Pol Data Structure"

POL_CODES = list(range(1, 7))        # Pol 1..6
POL_LABELS = [f"Pol {i}" for i in range(1, 7)]
# --- Consistent visual style per Pol (color, marker, linestyle) ---
POL_STYLES = {
    1: {"color": "#1f77b4", "marker": "o", "linestyle": "-"},   # blue circles
    2: {"color": "#ff7f0e", "marker": "s", "linestyle": "-"},   # orange squares
    3: {"color": "#2ca02c", "marker": "^", "linestyle": "-"},   # green triangles
    4: {"color": "#d62728", "marker": "D", "linestyle": "-"},   # red diamonds
    5: {"color": "#9467bd", "marker": "v", "linestyle": "-"},   # purple inverted triangle
    6: {"color": "#8c564b", "marker": "*", "linestyle": "-"},   # brown star
}
DEFAULT_POL_STYLES = {k: dict(v) for k, v in POL_STYLES.items()}  # for "Reset to defaults"

MARKER_OPTIONS = [
    ("Circle", "o"), ("Square", "s"), ("Triangle up", "^"), ("Diamond", "D"),
    ("Triangle down", "v"), ("Star", "*"), ("X", "x"), ("Plus", "+"),
    ("Pentagon", "p"), ("Hexagon", "h"), ("None", "None"),
]
LINESTYLE_OPTIONS = [
    ("Solid", "-"), ("Dashed", "--"), ("Dash-dot", "-."), ("Dotted", ":"), ("None", "None"),
]

# Marker cycle for "vary symbols per dataset" in overlays - assigned by first-appearance
# order of each key, independent of POL_STYLES.
OVERLAY_MARKER_CYCLE = ["o", "^", "s", "D", "v", "*", "P", "X", "p", "h"]

# Color cycle for "vary colours per dataset" in overlays - assigned by first-appearance
# order of each key, independent of POL_STYLES. When active, color no longer encodes Pol;
# Pol is then only distinguishable via the legend label ("<key> - Pol X").
OVERLAY_COLOR_CYCLE = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
]

# Statusbar timer id
_status_after_id = None


# ----------------------------
# Helpers
# ----------------------------
def normalize_str(x) -> str:
    """Safe string normalizer: cast to str, strip whitespace."""
    return "" if x is None else str(x).strip()

def ensure_openpyxl():
    """Ensure openpyxl is available; show a message if not."""
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        messagebox.showerror(
            "Missing dependency",
            "The 'openpyxl' package is required to read/write .xlsx files.\n\n"
            "Install it with:\n    pip install openpyxl"
        )
        return False

def interpolate_current(voltage_values, current_values, target):
    """Return current at a target voltage using exact match or linear interpolation."""
    try:
        s = pd.DataFrame({"V": voltage_values, "J": current_values}).dropna().sort_values("V")
        if s.empty:
            return None

        exact = s[s["V"] == target]
        if not exact.empty:
            return float(exact["J"].iloc[0])

        lower = s[s["V"] < target].tail(1)
        upper = s[s["V"] > target].head(1)
        if lower.empty or upper.empty:
            return None

        v1, j1 = float(lower["V"].iloc[0]), float(lower["J"].iloc[0])
        v2, j2 = float(upper["V"].iloc[0]), float(upper["J"].iloc[0])
        if v2 == v1:
            return None

        return j1 + (target - v1) * (j2 - j1) / (v2 - v1)
    except Exception:
        return None

def to_float_or_nan(v):
    """Return float(v) or np.nan for N/A-like values."""
    if v is None or (isinstance(v, float) and np.isnan(v)) or v == "N/A":
        return np.nan
    try:
        return float(v)
    except Exception:
        return np.nan

def round_sig_numeric(v, sig=4):
    """Round numbers to `sig` significant digits; preserve NaN for missing."""
    x = to_float_or_nan(v)
    if pd.isna(x) or x == 0:
        return 0.0 if x == 0 else np.nan
    power = sig - 1 - int(math.floor(math.log10(abs(x))))
    return round(x, power)

def parse_active_area(default_area=25.0) -> float:
    """
    Read the active area (cm²) from the GUI entry.
    Accepts comma/point decimals. Returns default_area if invalid.
    """
    try:
        s = active_area_var.get().strip()
        x = float(s.replace(",", "."))
        if x <= 0:
            raise ValueError
        return x
    except Exception:
        messagebox.showerror(
            "Invalid active area",
            "Please enter a positive number for the active area (cm²). Using 25 cm²."
        )
        return default_area

def _parse_optional_float(s: str):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return float(s.replace(",", "."))
    except Exception:
        return None

def get_iv_axes_limits():
    """
    Returns (xmin, xmax, ymin, ymax) for the Polarization curve (J/U). Empty fields => None (auto).
    """
    xmin = _parse_optional_float(iv_xmin_var.get())
    xmax = _parse_optional_float(iv_xmax_var.get())
    ymin = _parse_optional_float(iv_ymin_var.get())
    ymax = _parse_optional_float(iv_ymax_var.get())
    if xmin is not None and xmax is not None and xmin >= xmax:
        messagebox.showerror("Axis limits", "Polarization: J min must be < J max.")
        return None
    if ymin is not None and ymax is not None and ymin >= ymax:
        messagebox.showerror("Axis limits", "Polarization: U min must be < U max.")
        return None
    return xmin, xmax, ymin, ymax

def get_power_axes_limits():
    """
    Returns (xmin, xmax, ymin, ymax) for Power (J/P). Empty fields => None (auto).
    """
    xmin = _parse_optional_float(power_xmin_var.get())
    xmax = _parse_optional_float(power_xmax_var.get())
    ymin = _parse_optional_float(power_ymin_var.get())
    ymax = _parse_optional_float(power_ymax_var.get())
    if xmin is not None and xmax is not None and xmin >= xmax:
        messagebox.showerror("Axis limits", "Power: J min must be < J max.")
        return None
    if ymin is not None and ymax is not None and ymin >= ymax:
        messagebox.showerror("Axis limits", "Power: P min must be < P max.")
        return None
    return xmin, xmax, ymin, ymax

def fmt_axes(lims, xlab="x", ylab="y"):
    """Pretty short text for status bar."""
    if lims is None:
        return f"{xlab}[auto] {ylab}[auto]"
    xmin, xmax, ymin, ymax = lims
    def f(v):
        return "auto" if v is None else f"{float(v):.4g}"
    return f"{xlab}[{f(xmin)},{f(xmax)}] {ylab}[{f(ymin)},{f(ymax)}]"

def set_status(msg: str, timeout_ms: int = 4000):
    """Show a transient message in the status bar."""
    global _status_after_id
    try:
        status_var.set(msg)
        if _status_after_id is not None:
            root.after_cancel(_status_after_id)
            _status_after_id = None
        if timeout_ms and timeout_ms > 0:
            _status_after_id = root.after(timeout_ms, lambda: status_var.set("Ready"))
    except Exception:
        pass

def parse_block_size(default_bs=10) -> int:
    """
    Read 'points per average' from GUI. Must be positive integer.
    Falls back to default_bs if invalid.
    """
    try:
        bs = int(block_size_var.get().strip())
        if bs <= 0:
            raise ValueError
        return bs
    except Exception:
        messagebox.showerror("Averaging", f"Points per average must be a positive integer. Using {default_bs}.")
        return default_bs

def to_numeric_series(s):
    """Comma-to-dot & numeric conversion with coercion to NaN."""
    return pd.to_numeric(s.astype(str).str.replace(",", ".", regex=False), errors="coerce")


# ----------------------------
# Data scraping & cleanup
# ----------------------------
# Measurement files are named "<something>NN_YYYYMMDD.txt". NN is the run number within
# the folder: 01 is the intended measurement, 02/03/... are repeats or aborted attempts.
# The range 01-99 reproduces exactly what the two former patterns matched between them.
_FILE_SEQ_RE = re.compile(r"(0[1-9]|[1-9][0-9])_\d{8}\.txt$")
_PRIMARY_SEQ = 1


def _seq_label(n: int) -> str:
    return f"{n:02d}_"


def _has_pol_data(df) -> bool:
    """Does this frame hold rows extract_current() can actually use? Same test it applies
    itself (raw SetMarker vs POL_CODES), so a run declared usable here never turns out
    unusable there."""
    return "SetMarker" in df.columns and bool(df["SetMarker"].isin(POL_CODES).any())


def data_structure_status(key: str) -> str:
    """One sentence on where this dataset's numbers came from, and why they are missing
    when they are.

    Two independent findings feed it: what scrape_data() saw in the folder, and what
    extract_current() could make of it. The extraction note wins when there is one - a row
    without numbers raises the question "why is this empty?", and the answer is more useful
    there than a remark about an unused file.

    Unknown keys read "Missing" rather than blank: an empty cell looks like "not checked",
    which is the opposite of the truth.
    """
    note = pol_data_notes.get(key)
    if note:
        used = (pol_data_runs.get(key) or {}).get("used")
        where = f"{_seq_label(used)} used, " if used else ""
        return f"Check data! ({where}{note})"
    return pol_data_structure.get(key, "Missing")


def scrape_data(path: str) -> dict:
    """Read the measurement .txt files under `path` into one DataFrame per folder.

    Per folder only ONE run number is used: 01_ when it is there and readable, otherwise
    the lowest one that is (02_, then 03_, ...). Mixing run numbers would concatenate two
    separate measurements into a single curve, so the others are reported instead - both
    in the warning dialog and, permanently, in the pol_data_structure column that the
    results table and every export carry.
    """
    global pol_data_structure, pol_data_runs, pol_data_notes
    pol_data_structure = {}
    pol_data_runs = {}
    pol_data_notes = {}

    paths = glob.glob(os.path.join(path, "**/*.txt"), recursive=True)

    # folder -> {run number: [file paths]}
    candidates = {}
    for p in paths:
        m = _FILE_SEQ_RE.search(os.path.basename(p))
        if not m:
            continue
        folder_name = os.path.basename(os.path.dirname(p))
        candidates.setdefault(folder_name, {}).setdefault(int(m.group(1)), []).append(p)

    grouped = {}
    fallback_folders = []   # read something other than 01_
    extra_folders = []      # read 01_, but further run numbers lie unused
    missing_folders = []    # candidate files exist, none of them could be read

    for folder_name, by_seq in sorted(candidates.items()):
        # Pick the run to read, in two tiers. Preferred: the lowest run that holds rows
        # marked Pol 1-6, i.e. one the extraction can do something with. An aborted
        # measurement often leaves a handful of rows with no marker at all - not empty, so
        # a mere "has rows" test would park the folder on that stub while the real
        # measurement sits next to it as 02_.
        # Fallback tier: if NO run has markers, keep the first readable one anyway. The
        # folder then stays visible with its reason spelled out by extract_current(),
        # instead of silently vanishing from the dataset list.
        used_seq, frames = None, []
        weak_seq, weak_frames = None, []
        for seq in sorted(by_seq):
            attempt = []
            for p in sorted(by_seq[seq]):
                try:
                    df = pd.read_csv(p, encoding='cp1252', sep='\t',
                                     decimal=',', low_memory=False)
                except Exception as e:
                    print(f"Failed to read {p}: {e}", file=sys.stderr)
                    continue
                if not df.empty:
                    attempt.append(df)
                else:
                    print(f"Empty (header only): {p}", file=sys.stderr)
            if not attempt:
                continue
            if weak_seq is None:
                weak_seq, weak_frames = seq, attempt
            if any(_has_pol_data(df) for df in attempt):
                used_seq, frames = seq, attempt
                break
            print(f"No Pol 1-6 markers in {_seq_label(seq)} of '{folder_name}'", file=sys.stderr)

        if used_seq is None and weak_seq is not None:
            used_seq, frames = weak_seq, weak_frames

        if used_seq is None:
            pol_data_structure[folder_name] = "Missing"
            missing_folders.append(folder_name)
            continue

        try:
            grouped[folder_name] = pd.concat(frames, ignore_index=True)
        except Exception as e:
            print(f"Failed to concat for {folder_name}: {e}", file=sys.stderr)
            pol_data_structure[folder_name] = "Missing"
            missing_folders.append(folder_name)
            continue

        others = [s for s in sorted(by_seq) if s != used_seq]
        pol_data_runs[folder_name] = {"used": used_seq, "unused": others}
        unused = ", ".join(_seq_label(s) for s in others)

        if used_seq != _PRIMARY_SEQ:
            # Not "no 01_": it may well exist and simply be empty or unmarked. Say what
            # was read, which is the part that affects the numbers.
            pol_data_structure[folder_name] = f"Check data! (used {_seq_label(used_seq)}, not 01_)"
            fallback_folders.append(f"    {folder_name}  →  {_seq_label(used_seq)}")
        elif unused:
            # Starts with "OK" on purpose: the dataset itself is fine, 01_ was used as
            # intended. The remark is about a file left lying in the folder, not about
            # these numbers - reading it as a warning about the data was the whole
            # confusion with the earlier "Check data! (unused: 02_)".
            pol_data_structure[folder_name] = f"OK — 01_ used ({unused} ignored)"
            extra_folders.append(f"    {folder_name}  →  {unused}")
        else:
            pol_data_structure[folder_name] = "OK"

    # --- After scraping, report anything that was not a plain 01_-only folder ---
    if fallback_folders or extra_folders or missing_folders:
        lines = ["⚠ WARNING: Alternative Measurement Data Detected ⚠", ""]
        if fallback_folders:
            lines += ["01_ missing or empty — a later run was read INSTEAD:", ""] + fallback_folders + [""]
        if extra_folders:
            lines += ["01_ was read; these runs were NOT used:", ""] + extra_folders + [""]
        if missing_folders:
            lines += ["No readable measurement file at all:", ""] + \
                     [f"    {f}" for f in missing_folders] + [""]
        lines.append(f'Details stay visible in the "{DATA_STRUCTURE_COLUMN}" column')
        lines.append("of the results table and the exported file.")
        msg_body = "\n".join(lines)

        # Messagebox with yellow warning icon
        messagebox.showwarning("Alternative measurement data found", msg_body)
        print(msg_body, file=sys.stderr)

    return grouped


# ----------------------------
# Archives (.rar / .zip) as a data source
# ----------------------------
ARCHIVE_EXTS = (".rar", ".zip")
UNRAR_FILENAME = "UnRAR.exe"
# Keep a console window from flashing up out of the windowed (console=False) build.
_NO_WINDOW = getattr(subprocess, "CREATE_NO_WINDOW", 0)


def extract_archive(archive_path: str, dest_root: str) -> str:
    """Unpack an archive into <dest_root>/<archive name>/ and return that folder.

    The extra folder level is deliberate: scrape_data() keys datasets on the *parent
    folder name*, so .txt files lying loose at the archive root would otherwise be
    grouped under the random temp-dir name instead of the archive's name.
    Raises RuntimeError with a message meant for a messagebox.
    """
    ext = os.path.splitext(archive_path)[1].lower()
    target = os.path.join(dest_root, os.path.splitext(os.path.basename(archive_path))[0])
    os.makedirs(target, exist_ok=True)

    if ext == ".zip":
        try:
            with zipfile.ZipFile(archive_path) as zf:
                zf.extractall(target)   # stdlib sanitises member paths
        except Exception as e:
            raise RuntimeError(f"Could not unpack the ZIP archive:\n{e}")
        return target

    if ext == ".rar":
        tool = _find_file(UNRAR_FILENAME)
        if tool is None:
            raise RuntimeError(
                f"{UNRAR_FILENAME} was not found, so .rar archives cannot be opened.\n\n"
                f"Put {UNRAR_FILENAME} next to the program, or unpack the archive by hand "
                "and use the folder instead."
            )
        try:
            # x = extract with paths, -y = assume yes, -idq = quiet
            subprocess.run([tool, "x", "-y", "-idq", archive_path, target + os.sep],
                           check=True, capture_output=True, creationflags=_NO_WINDOW)
        except subprocess.CalledProcessError as e:
            detail = (e.stderr or e.stdout or b"").decode("cp1252", "replace").strip()
            raise RuntimeError(
                "Could not unpack the RAR archive.\n"
                "Password-protected and multi-part archives are not supported.\n\n"
                f"{detail[-400:]}"
            )
        return target

    raise RuntimeError(f"Unsupported archive type '{ext}'. Use {' or '.join(ARCHIVE_EXTS)}.")


# ----------------------------
# Extraction of Currents, OCV, and Pmax  (SetMarker-only)
# ----------------------------
def extract_current():
    """Compute i @ 0.6 V, i @ 0.65 V (A/cm²), OCV (V), and Pmax (W/cm²) per Pol 1..6 per folder."""
    global current_at_06V, pol_data_notes
    current_at_06V = {}
    # Recomputed from scratch on every run: the reasons depend on the active area and on
    # which data are loaded, both of which the user can change between runs.
    pol_data_notes = {}

    if not data_dict:
        messagebox.showinfo("Info", "No data available. Please load data first.")
        return

    for folder_name, df in data_dict.items():
        current_at_06V[folder_name] = {}

        # Required columns
        if "U Mittel [V]" not in df.columns or "I Mittel [A]" not in df.columns:
            print(f"Warning: Required columns not found in folder '{folder_name}'.", file=sys.stderr)
            pol_data_notes[folder_name] = "no U/I columns"
            continue

        # Normalize numeric columns (decimal comma → dot), convert current to A/cm² via active area
        try:
            area = parse_active_area()
            df = df.copy()
            df["U Mittel [V]"] = pd.to_numeric(
                df["U Mittel [V]"].astype(str).str.replace(",", ".", regex=False),
                errors="coerce"
            )
            df["I Mittel [A]"] = pd.to_numeric(
                df["I Mittel [A]"].astype(str).str.replace(",", ".", regex=False),
                errors="coerce"
            ) / area
        except Exception as e:
            print(f"Error processing folder '{folder_name}': {e}", file=sys.stderr)
            pol_data_notes[folder_name] = "U/I values unreadable"
            continue

        # Optional power-preferred columns
        has_U1 = "U1 Mittel [V]" in df.columns
        has_Jf = "JFilter [A/cm²]" in df.columns
        if has_U1:
            df["U1 Mittel [V]"] = pd.to_numeric(
                df["U1 Mittel [V]"].astype(str).str.replace(",", ".", regex=False),
                errors="coerce"
            )
        if has_Jf:
            df["JFilter [A/cm²]"] = pd.to_numeric(
                df["JFilter [A/cm²]"].astype(str).str.replace(",", ".", regex=False),
                errors="coerce"
            )

        # --- require SetMarker, no Kommentar fallback
        if "SetMarker" not in df.columns:
            print(f"Notice: '{folder_name}' has no SetMarker column. Skipping extraction for this key.", file=sys.stderr)
            pol_data_notes[folder_name] = "no SetMarker column"
            continue

        for pol_code in POL_CODES:
            pol_lbl = f"Pol {pol_code}"
            mask = df["SetMarker"] == pol_code

            filtered = df.loc[mask].dropna(subset=["U Mittel [V]", "I Mittel [A]"])
            if filtered.empty:
                print(f"Notice: No data for {pol_lbl} in '{folder_name}'.", file=sys.stderr)
                continue

            V = filtered["U Mittel [V]"]
            J = filtered["I Mittel [A]"]  # A/cm²

            # Currents at target voltages
            j_06  = interpolate_current(V, J, 0.6)
            j_065 = interpolate_current(V, J, 0.65)

            # OCV: voltage at |current| minimum
            try:
                idx = (J.abs()).idxmin()
                ocv = float(V.loc[idx]) if idx in V.index else None
            except Exception:
                ocv = None

            # Pmax: prefer U1/JFilter when available
            U_power = filtered["U1 Mittel [V]"] if has_U1 else V
            J_power = filtered["JFilter [A/cm²]"] if has_Jf else J
            power_density = (U_power * J_power).replace([np.inf, -np.inf], np.nan)
            pmax = float(power_density.max(skipna=True)) if power_density.notna().any() else None

            # Store consistent keys used by table & Excel updater
            current_at_06V[folder_name][f"{pol_lbl} i @ 0.6 V [A/cm²]"]  = j_06  if j_06  is not None else "N/A"
            current_at_06V[folder_name][f"{pol_lbl} i @ 0.65 V [A/cm²]"] = j_065 if j_065 is not None else "N/A"
            current_at_06V[folder_name][f"{pol_lbl} OCV [V]"]            = ocv   if ocv   is not None else "N/A"
            current_at_06V[folder_name][f"{pol_lbl} Pmax [W/cm²]"]       = pmax  if pmax  is not None else "N/A"

        # The loop can run to the end and still leave the row empty or half empty. All of
        # these look identical in the table - a lot of "N/A" - so name the actual cause.
        got = current_at_06V[folder_name]
        real = {k: v for k, v in got.items() if v != "N/A"}
        if not real:
            if not _has_pol_data(df):
                pol_data_notes[folder_name] = "no Pol 1-6 markers"
            elif not got:
                # Marked rows exist, but U or I could not be read as numbers in any of them.
                pol_data_notes[folder_name] = "Pol rows have no numeric U/I"
            else:
                pol_data_notes[folder_name] = "no usable values"
        elif not any(k.endswith("i @ 0.6 V [A/cm²]") for k in real):
            # OCV and Pmax come out of almost any curve; i @ 0.6 V does not. Losing the
            # headline metric while the row still looks populated is worth saying out loud.
            pol_data_notes[folder_name] = "no i @ 0.6 V in any Pol"

    show_current_table()
    set_status(f"Extracted metrics for {len(current_at_06V)} keys")


# ----------------------------
# Table (GUI) with i@0.6, i@0.65, OCV, Pmax
# ----------------------------
def show_current_table():
    if not current_at_06V:
        messagebox.showinfo("Info", "No data to display.")
        return

    def first_present(d: dict, keys):
        for k in keys:
            if k in d:
                return d[k]
        return "N/A"

    def fmt(x):
        try:
            if x is None or (isinstance(x, float) and np.isnan(x)) or x == "N/A":
                return "N/A"
            return f"{float(x):.4g}"  # compact numeric formatting (4 sig figs)
        except Exception:
            return str(x)

    table_window = tk.Toplevel(root)
    table_window.title("Currents, OCV, Pmax per Pol")

    cols = ("Key", "Pol", "i @ 0.6 V [A/cm²]", "i @ 0.65 V [A/cm²]", "OCV [V]", "Pmax [W/cm²]",
            DATA_STRUCTURE_COLUMN)
    tree = ttk.Treeview(table_window, columns=cols, show="headings")

    widths = {"Key": 180, DATA_STRUCTURE_COLUMN: 220}
    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, width=widths.get(c, 160), anchor="center")

    tree.pack(fill="both", expand=True)

    for key, values in current_at_06V.items():
        structure = data_structure_status(key)
        for pol_lbl in POL_LABELS:
            i06 = first_present(values, [
                f"{pol_lbl} i @ 0.6 V [A/cm²]",
                f"{pol_lbl.replace(' ', '')} i @ 0.6 V [A/cm²]",
                f"{pol_lbl.replace(' ', '')} Current @ 0.6V [A]"
            ])
            i065 = first_present(values, [
                f"{pol_lbl} i @ 0.65 V [A/cm²]",
                f"{pol_lbl.replace(' ', '')} i @ 0.65 V [A/cm²]"
            ])
            ocv = first_present(values, [f"{pol_lbl} OCV [V]", f"{pol_lbl.replace(' ', '')} OCV [V]"])
            pmax = first_present(values, [
                f"{pol_lbl} Pmax [W/cm²]",
                f"{pol_lbl.replace(' ', '')} Pmax [W/cm²]"
            ])

            tree.insert("", "end",
                        values=(key, pol_lbl, fmt(i06), fmt(i065), fmt(ocv), fmt(pmax), structure))


# ----------------------------
# Plotting (SetMarker-only)
# ----------------------------

def plot_data(folder_name, df):
    """Plot grouped mean Polarization curves (by N-point blocks) for Pol1..Pol6."""
    fig, ax = plt.subplots(figsize=(10, 6))

    area = parse_active_area()
    bs = parse_block_size()

    if "SetMarker" not in df.columns:
        messagebox.showwarning("Data", "This dataset has no 'SetMarker' column. Cannot plot by Pol.")
        return

    for pol_code in POL_CODES:

        pol_lbl = f"Pol {pol_code}"
        filtered = df[df["SetMarker"] == pol_code].copy()
        if filtered.empty:
            continue

        # Polarization
        filtered["U Mittel [V]"] = to_numeric_series(filtered["U Mittel [V]"])
        filtered["I Mittel [A]"] = to_numeric_series(filtered["I Mittel [A]"]) / area
        filtered_iv = filtered.dropna(subset=["U Mittel [V]", "I Mittel [A]"])
        grouped_iv = filtered_iv.groupby(np.arange(len(filtered_iv)) // bs).agg({
            "U Mittel [V]": "mean",
            "I Mittel [A]": "mean"
        })
        if not grouped_iv.empty:
            style = POL_STYLES.get(pol_code, {"color": None, "marker": "o", "linestyle": "-"})
            ax.plot(grouped_iv["I Mittel [A]"], grouped_iv["U Mittel [V]"],
                    color=style["color"], marker=style["marker"], linestyle=style["linestyle"], label=pol_lbl)

    ax.set_title(f"Polarization Curves for {folder_name}")
    ax.set_xlabel("Current density J [A/cm²]")
    ax.set_ylabel("Voltage U [V]")
    ax.grid(True)


    # Legends
    handles_iv, labels_iv = ax.get_legend_handles_labels()
    if handles_iv:
        ax.legend(handles_iv, labels_iv, title="Pol (Polarization)", loc="upper right", frameon=True)
        leg = ax.legend(loc="upper left", frameon=True)
        leg.set_draggable(True)
        leg.get_frame().set_alpha(0.9)  # slight opacity so lines are still visible
        leg.get_frame().set_facecolor("white")
        leg.get_frame().set_edgecolor("black")

    # apply manual Polarization limits if provided
    lims = get_iv_axes_limits()
    if lims is not None:
        xmin, xmax, ymin, ymax = lims
        if xmin is not None or xmax is not None:
            ax.set_xlim(left=xmin, right=xmax)
        if ymin is not None or ymax is not None:
            ax.set_ylim(bottom=ymin, top=ymax)

    set_status(f"Polarization {folder_name} | area={area:g} cm² | bs={bs} | {fmt_axes(lims, 'J','U')}")
    attach_label_editor(fig, ax)
    plt.show()

def plot_power_data(folder_name, df):
    """Plot grouped mean power curves (P vs J) for Pol1..Pol6 with optional Pmax table (lower-right)."""
    fig, ax = plt.subplots(figsize=(10.5, 6))

    area = parse_active_area()
    bs = parse_block_size()


    if "SetMarker" not in df.columns:
        messagebox.showwarning("Data", "This dataset has no 'SetMarker' column. Cannot plot by Pol.")
        return

    pmax_rows = []  # (pol_code, pol_label, J*, Pmax)

    for pol_code in POL_CODES:
        pol_lbl = f"Pol {pol_code}"

        filtered = df[df["SetMarker"] == pol_code].copy()
        if filtered.empty:
            continue

        # Voltage source: prefer U1 Mittel [V] if available
        if "U1 Mittel [V]" in filtered.columns:
            filtered["U_plot"] = to_numeric_series(filtered["U1 Mittel [V]"])
        else:
            filtered["U_plot"] = to_numeric_series(filtered["U Mittel [V]"])

        # Current density: prefer JFilter [A/cm²], else I Mittel [A] / area
        if "JFilter [A/cm²]" in filtered.columns:
            filtered["J_plot"] = to_numeric_series(filtered["JFilter [A/cm²]"])
        else:
            filtered["J_plot"] = to_numeric_series(filtered["I Mittel [A]"]) / area

        filtered = filtered.dropna(subset=["U_plot", "J_plot"])
        if filtered.empty:
            continue

        # N-point block averaging (same as Polarization)
        grouped = filtered.groupby(np.arange(len(filtered)) // bs).agg({
            "U_plot": "mean",
            "J_plot": "mean"
        })
        if grouped.empty:
            continue

        # Power density
        grouped["P [W/cm²]"] = grouped["U_plot"] * grouped["J_plot"]

        # Plot P vs J
        style = POL_STYLES.get(pol_code, {"color": None, "marker": "o", "linestyle": "-"})
        ax.plot(grouped["J_plot"], grouped["P [W/cm²]"],
                color=style["color"], marker=style["marker"], linestyle=style["linestyle"], label=pol_lbl)

        # Collect Pmax for table
        if grouped["P [W/cm²]"].notna().any():
            imax = grouped["P [W/cm²]"].idxmax()
            j_star = grouped.loc[imax, "J_plot"]
            p_star = grouped.loc[imax, "P [W/cm²]"]
            if np.isfinite(j_star) and np.isfinite(p_star):
                pmax_rows.append((pol_code, pol_lbl, j_star, p_star))

    # Cosmetics
    ax.set_title(f"Power Curves for {folder_name}")
    ax.set_xlabel("Current density J [A/cm²]")
    ax.set_ylabel("Power density P [W/cm²]")
    ax.grid(True)

    # Legend INSIDE, upper-right
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(handles, labels, title="Pol", loc="upper right", frameon=True)
        leg = ax.legend(loc="upper left", frameon=True)
        leg.set_draggable(True)
        leg.get_frame().set_alpha(0.9)
        leg.get_frame().set_facecolor("white")
        leg.get_frame().set_edgecolor("black")

    # apply manual Power axis limits if provided
    plims = get_power_axes_limits()
    if plims is not None:
        xmin, xmax, ymin, ymax = plims
        if xmin is not None or xmax is not None:
            ax.set_xlim(left=xmin, right=xmax)
        if ymin is not None or ymax is not None:
            ax.set_ylim(bottom=ymin, top=ymax)

    # Pmax table INSIDE, lower-right (only if checkbox is on)
    if 'show_pmax_var' in globals() and show_pmax_var.get() and pmax_rows:
        pmax_rows.sort(key=lambda t: t[0])
        table_data = [[lbl, f"{j:.4g}", f"{p:.4g}"] for _, lbl, j, p in pmax_rows]
        col_labels = ["Pol", "J* [A/cm²]", "Pmax [W/cm²]"]

        n = len(table_data)
        height = min(0.40, 0.10 + 0.05 * n)
        bbox = [0.62, 0.02, 0.36, height]  # [left, bottom, width, height] in AXES coords

        # OPAQUE WHITE BACKGROUND PANEL behind the table
        pad = 0.01
        bg = Rectangle(
            (bbox[0] - pad, bbox[1] - pad),
            bbox[2] + 2*pad, bbox[3] + 2*pad,
            transform=ax.transAxes,
            facecolor="white", edgecolor="black", linewidth=0.8,
            zorder=5, clip_on=False
        )
        ax.add_patch(bg)

        # Table on top of the panel
        tbl = ax.table(
            cellText=table_data,
            colLabels=col_labels,
            loc="lower right",
            bbox=bbox
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(12)
        tbl.scale(1.0, 1.08)
        tbl.set_zorder(6)

        for cell in tbl.get_celld().values():
            cell.set_facecolor("white")
            cell.set_edgecolor("black")
            cell.set_alpha(1.0)
            cell.set_zorder(7)

        ax.set_axisbelow(True)

    set_status(f"Power {folder_name} | area={area:g} cm² | bs={bs} | {fmt_axes(plims, 'J','P')}")
    attach_label_editor(fig, ax)
    plt.show()

# --- Inline editor for axes/legend text (press 'E' in the figure window) ---
import tkinter.simpledialog as simpledialog

def _open_label_editor(fig, ax, ax2=None):
    # Current values
    cur_title = ax.get_title() or ""
    cur_xlabel = ax.get_xlabel() or ""
    cur_ylabel = ax.get_ylabel() or ""
    cur_ylabel_right = ax2.get_ylabel() if ax2 is not None else ""

    leg_left = ax.get_legend()
    leg_right = ax2.get_legend() if ax2 is not None else None

    cur_leg_title = (leg_left.get_title().get_text() if leg_left else "")
    # Legend item labels (prefer left legend; if none, use right)
    legend_source = leg_left or leg_right
    cur_items = []
    if legend_source:
        cur_items = [t.get_text() for t in legend_source.get_texts()]

    # Dialog
    top = tk.Toplevel()
    top.title("Edit labels")
    top.grab_set()

    def add_row(r, text, var, width=48):
        tk.Label(top, text=text).grid(row=r, column=0, sticky="e", padx=6, pady=3)
        e = tk.Entry(top, textvariable=var, width=width, justify="left")
        e.grid(row=r, column=1, sticky="w", padx=6, pady=3)
        return e

    v_title = tk.StringVar(value=cur_title)
    v_xlabel = tk.StringVar(value=cur_xlabel)
    v_ylabel = tk.StringVar(value=cur_ylabel)
    v_ylabel_r = tk.StringVar(value=cur_ylabel_right)
    v_leg_title = tk.StringVar(value=cur_leg_title)

    r = 0
    add_row(r, "Title:", v_title); r += 1
    add_row(r, "X label:", v_xlabel); r += 1
    add_row(r, "Y label (left):", v_ylabel); r += 1
    if ax2 is not None:
        add_row(r, "Y label (right):", v_ylabel_r); r += 1
    add_row(r, "Legend title:", v_leg_title); r += 1

    tk.Label(top, text="Legend items (one per line):").grid(row=r, column=0, sticky="ne", padx=6, pady=3)
    txt_items = tk.Text(top, width=48, height=max(6, len(cur_items) or 6))
    txt_items.grid(row=r, column=1, sticky="w", padx=6, pady=3)
    if cur_items:
        txt_items.insert("1.0", "\n".join(cur_items))
    r += 1

    btns = tk.Frame(top); btns.grid(row=r, column=0, columnspan=2, pady=8)
    def apply_and_close():
        # Axes titles/labels
        ax.set_title(v_title.get())
        ax.set_xlabel(v_xlabel.get())
        ax.set_ylabel(v_ylabel.get())
        if ax2 is not None:
            ax2.set_ylabel(v_ylabel_r.get())

        # Legend title + items (left first, else right)
        target_leg = ax.get_legend() or (ax2.get_legend() if ax2 is not None else None)
        if target_leg:
            # title
            target_leg.set_title(v_leg_title.get())
            # items
            new_labels = [line.strip() for line in txt_items.get("1.0", "end").splitlines()]
            texts = target_leg.get_texts()
            # Resize if needed: only rename up to min of both
            for t, newt in zip(texts, new_labels):
                if newt:
                    t.set_text(newt)

        fig.canvas.draw_idle()
        top.destroy()

    tk.Button(btns, text="Apply", command=apply_and_close).pack(side="left", padx=6)
    tk.Button(btns, text="Cancel", command=top.destroy).pack(side="left", padx=6)

def attach_label_editor(fig, ax, ax2=None):
    """Press 'E' in the Matplotliedit labels/legend."""
    def _on_key(event):
        if event.key and event.key.lower() == "e":
            _open_label_editor(fig, ax, ax2)
    fig.canvas.mpl_connect("key_press_event", _on_key)


# ----------------------------
# Overlay (Key + Pol) — Polarization and Power
# ----------------------------
def overlay_iv_pairs(pairs, save=False):
    """
    Overlay Polarization curves for a list of (key, pol_code) pairs.
    Label each curve as '<key> – Pol X'.
    """
    if not pairs:
        messagebox.showwarning("Overlay", "No key/pol pairs selected.")
        return

    fig, ax = plt.subplots(figsize=(10, 6))

    area = parse_active_area()
    bs = parse_block_size()

    # Color and marker follow the Pol by default (consistent with every other plot type).
    # "Vary symbols per dataset" gives each key its own fixed marker instead, and
    # "Vary colours per dataset" gives each key its own fixed color instead, so curves
    # sharing a Pol across datasets stay distinguishable. When both are on, Pol is only
    # distinguishable via the legend label ("<key> - Pol X").
    vary_marker_by_dataset = 'vary_marker_by_dataset_var' in globals() and vary_marker_by_dataset_var.get()
    vary_color_by_dataset = 'vary_color_by_dataset_var' in globals() and vary_color_by_dataset_var.get()
    key_marker = {}
    key_color = {}
    for key, _ in pairs:
        if key not in key_marker:
            key_marker[key] = OVERLAY_MARKER_CYCLE[len(key_marker) % len(OVERLAY_MARKER_CYCLE)]
        if key not in key_color:
            key_color[key] = OVERLAY_COLOR_CYCLE[len(key_color) % len(OVERLAY_COLOR_CYCLE)]

    for key, pol_code in pairs:
        df = data_dict.get(key)
        if df is None or df.empty:
            print(f"[overlay_iv] Missing or empty data for key '{key}'", file=sys.stderr)
            continue
        if "SetMarker" not in df.columns:
            print(f"[overlay_iv] No SetMarker in key '{key}', skipping.", file=sys.stderr)
            continue

        filtered = df[df["SetMarker"] == pol_code].copy()
        if filtered.empty:
            print(f"[overlay_iv] No rows for {key} Pol {pol_code}", file=sys.stderr)
            continue

        # Polarization
        filtered["U Mittel [V]"] = to_numeric_series(filtered["U Mittel [V]"])
        filtered["I Mittel [A]"] = to_numeric_series(filtered["I Mittel [A]"]) / area
        filtered_iv = filtered.dropna(subset=["U Mittel [V]", "I Mittel [A]"])
        grouped_iv = filtered_iv.groupby(np.arange(len(filtered_iv)) // bs).agg({
            "U Mittel [V]": "mean",
            "I Mittel [A]": "mean"
        })
        if not grouped_iv.empty:
            style = POL_STYLES.get(pol_code, {"color": None, "marker": "o", "linestyle": "-"})
            marker = key_marker[key] if vary_marker_by_dataset else style["marker"]
            color = key_color[key] if vary_color_by_dataset else style["color"]
            ax.plot(
                grouped_iv["I Mittel [A]"], grouped_iv["U Mittel [V]"],
                color=color, marker=marker, linestyle=style["linestyle"],
                label=f"{key} – Pol {pol_code}"
            )

    ax.set_title(f"Polarization Overlay ({len(pairs)} curves)")
    ax.set_xlabel("Current density J [A/cm²]")
    ax.set_ylabel("Voltage U [V]")
    ax.grid(True)

    # Legend
    handles_iv, labels_iv = ax.get_legend_handles_labels()
    if handles_iv:
        ax.legend(handles_iv, labels_iv, title="Curves (Polarization)", loc="upper left", frameon=True)
        leg1 = ax.legend(handles_iv, labels_iv, loc="upper left", frameon=True)
        leg1.set_draggable(True)
        leg1.get_frame().set_alpha(0.9)
        leg1.get_frame().set_facecolor("white")
        leg1.get_frame().set_edgecolor("black")

    # Apply manual Polarization limits if set
    lims = get_iv_axes_limits()
    if lims is not None:
        xmin, xmax, ymin, ymax = lims
        if xmin is not None or xmax is not None:
            ax.set_xlim(left=xmin, right=xmax)
        if ymin is not None or ymax is not None:
            ax.set_ylim(bottom=ymin, top=ymax)

    set_status(
        f"Overlay Polarization | curves={len(pairs)} | bs={bs} | {fmt_axes(lims, 'J','U')}"
    )

    if save:
        outdir = filedialog.askdirectory(title="Choose output folder for overlay Polarization PNG")
        if not outdir:
            plt.close(fig); return
        path = os.path.join(outdir, "overlay_Polarization_pairs.png")
        fig.savefig(path, dpi=150, bbox_inches="tight")
        messagebox.showinfo("Saved", path)
        plt.close(fig)
    else:
        attach_label_editor(fig, ax)
        plt.show()


def overlay_power_pairs(pairs, save=False):
    """
    Overlay Power (P–J) curves for a list of (key, pol_code) pairs.
    Label each curve as '<key> – Pol X'.
    """
    if not pairs:
        messagebox.showwarning("Overlay", "No key/pol pairs selected.")
        return

    fig, ax = plt.subplots(figsize=(10, 6))

    area = parse_active_area()
    bs = parse_block_size()

    # Color and marker follow the Pol by default (consistent with every other plot type).
    # "Vary symbols per dataset" gives each key its own fixed marker instead, and
    # "Vary colours per dataset" gives each key its own fixed color instead, so curves
    # sharing a Pol across datasets stay distinguishable. When both are on, Pol is only
    # distinguishable via the legend label ("<key> - Pol X").
    vary_marker_by_dataset = 'vary_marker_by_dataset_var' in globals() and vary_marker_by_dataset_var.get()
    vary_color_by_dataset = 'vary_color_by_dataset_var' in globals() and vary_color_by_dataset_var.get()
    key_marker = {}
    key_color = {}
    for key, _ in pairs:
        if key not in key_marker:
            key_marker[key] = OVERLAY_MARKER_CYCLE[len(key_marker) % len(OVERLAY_MARKER_CYCLE)]
        if key not in key_color:
            key_color[key] = OVERLAY_COLOR_CYCLE[len(key_color) % len(OVERLAY_COLOR_CYCLE)]

    for key, pol_code in pairs:
        df = data_dict.get(key)
        if df is None or df.empty:
            print(f"[overlay_power] Missing or empty data for key '{key}'", file=sys.stderr)
            continue
        if "SetMarker" not in df.columns:
            print(f"[overlay_power] No SetMarker in key '{key}', skipping.", file=sys.stderr)
            continue

        filtered = df[df["SetMarker"] == pol_code].copy()
        if filtered.empty:
            print(f"[overlay_power] No rows for {key} Pol {pol_code}", file=sys.stderr)
            continue

        if "U1 Mittel [V]" in filtered.columns:
            filtered["U_plot"] = to_numeric_series(filtered["U1 Mittel [V]"])
        else:
            filtered["U_plot"] = to_numeric_series(filtered["U Mittel [V]"])

        if "JFilter [A/cm²]" in filtered.columns:
            filtered["J_plot"] = to_numeric_series(filtered["JFilter [A/cm²]"])
        else:
            filtered["J_plot"] = to_numeric_series(filtered["I Mittel [A]"]) / area

        filtered = filtered.dropna(subset=["U_plot", "J_plot"])
        if filtered.empty:
            continue

        grouped = filtered.groupby(np.arange(len(filtered)) // bs).agg({
            "U_plot": "mean", "J_plot": "mean"
        })
        if grouped.empty:
            continue

        grouped["P [W/cm²]"] = grouped["U_plot"] * grouped["J_plot"]
        style = POL_STYLES.get(pol_code, {"color": None, "marker": "o", "linestyle": "-"})
        marker = key_marker[key] if vary_marker_by_dataset else style["marker"]
        color = key_color[key] if vary_color_by_dataset else style["color"]
        ax.plot(
            grouped["J_plot"], grouped["P [W/cm²]"],
            color=color, marker=marker, linestyle=style["linestyle"],
            label=f"{key} – Pol {pol_code}"
        )

    ax.set_title(f"Power Overlay ({len(pairs)} curves)")
    ax.set_xlabel("Current density J [A/cm²]")
    ax.set_ylabel("Power density P [W/cm²]")
    ax.grid(True)
    leg = ax.legend(loc="upper left", frameon=True)
    leg.set_draggable(True)
    leg.get_frame().set_alpha(0.9)
    leg.get_frame().set_facecolor("white")
    leg.get_frame().set_edgecolor("black")



    # apply manual Power limits if given
    plims = get_power_axes_limits()
    if plims is not None:
        xmin, xmax, ymin, ymax = plims
        if xmin is not None or xmax is not None:
            ax.set_xlim(left=xmin, right=xmax)
        if ymin is not None or ymax is not None:
            ax.set_ylim(bottom=ymin, top=ymax)

    set_status(f"Overlay Power | curves={len(pairs)} | bs={bs} | {fmt_axes(plims, 'J','P')}")

    if save:
        outdir = filedialog.askdirectory(title="Choose output folder for overlay Power PNG")
        if not outdir:
            plt.close(fig); return
        path = os.path.join(outdir, "overlay_Power_pairs.png")
        fig.savefig(path, dpi=150, bbox_inches="tight")
        messagebox.showinfo("Saved", path)
        plt.close(fig)
    else:
        attach_label_editor(fig, ax)
        plt.show()


def open_overlay_keypol_selector():
    """
    Dialog to build a list of (key, pol) pairs and plot overlays.
    """
    if not data_dict:
        messagebox.showwarning("Overlay", "Please load data first.")
        return

    win = tk.Toplevel(root)
    win.title("Overlay: choose Key + Pol")
    win.grab_set()

    # Left: keys list
    tk.Label(win, text="Keys").grid(row=0, column=0, padx=8, pady=(8,2), sticky="w")
    lb_keys = tk.Listbox(win, selectmode="browse", width=36, height=12)
    for k in sorted(data_dict.keys()):
        lb_keys.insert("end", k)
    lb_keys.grid(row=1, column=0, rowspan=4, padx=8, pady=4, sticky="nsw")

    # Middle: combobox for Pol
    tk.Label(win, text="Pol").grid(row=0, column=1, padx=8, pady=(8,2), sticky="w")
    pol_choice = tk.StringVar(value="1")
    cb_pol = ttk.Combobox(win, textvariable=pol_choice, values=[str(i) for i in POL_CODES],
                          state="readonly", width=6)
    cb_pol.grid(row=1, column=1, padx=8, pady=2, sticky="nw")

    # Right: selected pairs
    tk.Label(win, text="Selected pairs").grid(row=0, column=2, padx=8, pady=(8,2), sticky="w")
    lb_pairs = tk.Listbox(win, selectmode="extended", width=42, height=12)
    lb_pairs.grid(row=1, column=2, rowspan=4, padx=8, pady=4, sticky="nsew")

    # Buttons: add/remove/clear
    def add_pair():
        sel = lb_keys.curselection()
        if not sel:
            messagebox.showwarning("Overlay", "Select a key on the left.")
            return
        key = lb_keys.get(sel[0])
        try:
            pol = int(pol_choice.get())
        except Exception:
            pol = 1
        if pol not in POL_CODES:
            messagebox.showwarning("Overlay", "Invalid Pol.")
            return
        df = data_dict.get(key)
        if "SetMarker" not in df.columns:
            messagebox.showwarning("Overlay", f"Key '{key}' has no SetMarker. Skipping.")
            return
        lb_pairs.insert("end", f"{key} ::: Pol {pol}")

    def remove_selected():
        for i in reversed(lb_pairs.curselection()):
            lb_pairs.delete(i)

    def clear_all():
        lb_pairs.delete(0, "end")

    tk.Button(win, text="Add", command=add_pair).grid(row=2, column=1, padx=8, pady=2, sticky="nw")
    tk.Button(win, text="Remove", command=remove_selected).grid(row=3, column=1, padx=8, pady=2, sticky="nw")
    tk.Button(win, text="Clear", command=clear_all).grid(row=4, column=1, padx=8, pady=2, sticky="nw")

    # Build pairs list from UI
    def collect_pairs():
        pairs = []
        for i in range(lb_pairs.size()):
            item = lb_pairs.get(i)
            # format: "<key> ::: Pol X"
            try:
                key, right = item.split(":::")
                key = key.strip()
                pol = int(right.strip().replace("Pol", "").strip())
                pairs.append((key, pol))
            except Exception:
                continue
        return pairs

    # Marker mode: same symbol per Pol (default) or a distinct symbol per dataset
    tk.Checkbutton(
        win, text="Vary symbols per dataset (within same Pol)",
        variable=vary_marker_by_dataset_var
    ).grid(row=5, column=0, columnspan=3, padx=8, pady=(4, 0), sticky="w")

    # Color mode: same color per Pol (default) or a distinct color per dataset
    tk.Checkbutton(
        win, text="Vary colours per dataset (within same Pol)",
        variable=vary_color_by_dataset_var
    ).grid(row=6, column=0, columnspan=3, padx=8, pady=(0, 0), sticky="w")

    # Bottom buttons: plot/save, Polarization/Power
    btns = tk.Frame(win); btns.grid(row=7, column=0, columnspan=3, pady=8)

    tk.Button(btns, text="Plot Polarization", command=lambda: overlay_iv_pairs(collect_pairs(), save=False)).pack(side="left", padx=8)
    tk.Button(btns, text="Plot Power", command=lambda: overlay_power_pairs(collect_pairs(), save=False)).pack(side="left", padx=8)
    tk.Button(btns, text="Save Polarization PNG", command=lambda: overlay_iv_pairs(collect_pairs(), save=True)).pack(side="left", padx=8)
    tk.Button(btns, text="Save Power PNG", command=lambda: overlay_power_pairs(collect_pairs(), save=True)).pack(side="left", padx=8)
    tk.Button(btns, text="Close", command=win.destroy).pack(side="right", padx=8)


# ----------------------------
# Export Results (CSV/XLSX) — Wide + Long, 4 sig figs
# ----------------------------
def export_results():
    """Export:
       • Wide + Long summary of extracted points (4 sig figs), with Wide reordered by metric blocks.
       • Averaged Polarization curve points (U,V) per Key & Pol, as a separate sheet/CSV.
    """
    if not current_at_06V:
        messagebox.showinfo("Info", "No results to export. Run 'Extract i@0.6 / i@0.65 + OCV + Pmax' first.")
        return

    # Helpers (reuse same fallback logic as the table)
    def first_present(d: dict, keys):
        for k in keys:
            if k in d:
                return d[k]
        return "N/A"

    # --- Desired metric block order for WIDE ---
    METRICS_ORDERED = [
        "i @ 0.6 V [A/cm²]",
        "i @ 0.65 V [A/cm²]",
        "OCV [V]",
        "Pmax [W/cm²]",
    ]
    # full ordered column list (excluding "Key")
    WIDE_COLS_ORDER = [f"Pol {i} {m}" for m in METRICS_ORDERED for i in range(1, 7)]

    # ---------- Build WIDE ----------
    wide_rows = []
    for key, values in current_at_06V.items():
        row = {"Key": key, DATA_STRUCTURE_COLUMN: data_structure_status(key)}
        for pol in range(1, 7):
            pol_lbl = f"Pol {pol}"
            # 0.6 V
            row[f"{pol_lbl} i @ 0.6 V [A/cm²]"] = first_present(values, [
                f"{pol_lbl} i @ 0.6 V [A/cm²]",
                f"{pol_lbl.replace(' ', '')} i @ 0.6 V [A/cm²]",
                f"{pol_lbl.replace(' ', '')} Current @ 0.6V [A]"
            ])
            # 0.65 V
            row[f"{pol_lbl} i @ 0.65 V [A/cm²]"] = first_present(values, [
                f"{pol_lbl} i @ 0.65 V [A/cm²]",
                f"{pol_lbl.replace(' ', '')} i @ 0.65 V [A/cm²]"
            ])
            # OCV
            row[f"{pol_lbl} OCV [V]"] = first_present(values, [
                f"{pol_lbl} OCV [V]",
                f"{pol_lbl.replace(' ', '')} OCV [V]"
            ])
            # Pmax
            row[f"{pol_lbl} Pmax [W/cm²]"] = first_present(values, [
                f"{pol_lbl} Pmax [W/cm²]",
                f"{pol_lbl.replace(' ', '')} Pmax [W/cm²]"
            ])
        wide_rows.append(row)

    df_wide = pd.DataFrame(wide_rows)

    # Ensure all ordered columns exist, then reorder: Key + (metric blocks)
    for c in WIDE_COLS_ORDER:
        if c not in df_wide.columns:
            df_wide[c] = np.nan
    # Structure verdict sits next to the Key, not at the far right behind 24 metric columns
    # where nobody scrolls to.
    df_wide = df_wide[["Key", DATA_STRUCTURE_COLUMN] + WIDE_COLS_ORDER]

    # ---------- Build LONG ----------
    long_rows = []
    for key, values in current_at_06V.items():
        for pol in range(1, 7):
            pol_lbl = f"Pol {pol}"
            long_rows.append({
                "Key": key,
                DATA_STRUCTURE_COLUMN: data_structure_status(key),
                "Pol": pol_lbl,
                "i @ 0.6 V [A/cm²]": first_present(values, [
                    f"{pol_lbl} i @ 0.6 V [A/cm²]",
                    f"{pol_lbl.replace(' ', '')} i @ 0.6 V [A/cm²]",
                    f"{pol_lbl.replace(' ', '')} Current @ 0.6V [A]"
                ]),
                "i @ 0.65 V [A/cm²]": first_present(values, [
                    f"{pol_lbl} i @ 0.65 V [A/cm²]",
                    f"{pol_lbl.replace(' ', '')} i @ 0.65 V [A/cm²]"
                ]),
                "OCV [V]": first_present(values, [
                    f"{pol_lbl} OCV [V]",
                    f"{pol_lbl.replace(' ', '')} OCV [V]"
                ]),
                "Pmax [W/cm²]": first_present(values, [
                    f"{pol_lbl} Pmax [W/cm²]",
                    f"{pol_lbl.replace(' ', '')} Pmax [W/cm²]"
                ]),
            })
    df_long = pd.DataFrame(long_rows)

    # ---------- Build Curves_Polarization (averaged Polarization points used in plots) ----------
    curves_rows = []
    area = parse_active_area()
    bs = parse_block_size()

    if not data_dict:
        print("[export] No raw data_dict available; skipping Curves_Polarization export.", file=sys.stderr)

    for key, df in (data_dict or {}).items():
        if "SetMarker" not in df.columns or df.empty:
            continue

        # Base conversions once to avoid repetition
        df_num = df.copy()
        df_num["U Mittel [V]"] = to_numeric_series(df_num["U Mittel [V]"])
        df_num["I Mittel [A]"] = to_numeric_series(df_num["I Mittel [A]"])

        for pol in range(1, 7):
            pol_lbl = f"Pol {pol}"
            fpol = df_num[df_num["SetMarker"] == pol].dropna(subset=["U Mittel [V]", "I Mittel [A]"])
            if fpol.empty:
                continue

            # J and U
            fpol["J [A/cm²]"] = fpol["I Mittel [A]"] / area
            fpol["U [V]"] = fpol["U Mittel [V]"]
            fpol = fpol.dropna(subset=["J [A/cm²]", "U [V]"])
            if fpol.empty:
                continue

            gidx = np.arange(len(fpol)) // bs
            g = fpol.groupby(gidx).agg({"U [V]": "mean", "J [A/cm²]": "mean"}).reset_index(drop=True)
            if g.empty:
                continue

            # round to 4 sig figs
            g["U [V]"] = g["U [V]"].apply(round_sig_numeric)
            g["J [A/cm²]"] = g["J [A/cm²]"].apply(round_sig_numeric)

            for _, r in g.iterrows():
                curves_rows.append({
                    "Key": key,
                    "Pol": pol_lbl,
                    "U [V]": r["U [V]"],
                    "J [A/cm²]": r["J [A/cm²]"],
                })

    df_curves = pd.DataFrame(curves_rows, columns=["Key", "Pol", "U [V]", "J [A/cm²]"])

    # ---------- Round Wide/Long numeric to 4 sig figs ----------
    # Text columns must stay out: round_sig_numeric() turns anything non-numeric into NaN,
    # which would silently erase the structure verdict.
    for c in [c for c in df_wide.columns if c not in ("Key", DATA_STRUCTURE_COLUMN)]:
        df_wide[c] = df_wide[c].apply(round_sig_numeric)
    for c in ["i @ 0.6 V [A/cm²]", "i @ 0.65 V [A/cm²]", "OCV [V]", "Pmax [W/cm²]"]:
        if c in df_long.columns:
            df_long[c] = df_long[c].apply(round_sig_numeric)

    # ---------- Save ----------
    save_path = filedialog.asksaveasfilename(
        title="Save results as...",
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx"), ("CSV (three files: _wide/_long/_curves_polarization)", "*.csv")]
    )
    if not save_path:
        return

    base, ext = os.path.splitext(save_path)
    ext = ext.lower()

    try:
        if ext == ".xlsx":
            if not ensure_openpyxl():
                return
            from openpyxl.utils import get_column_letter
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                df_wide.to_excel(writer, index=False, sheet_name="Wide")
                df_long.to_excel(writer, index=False, sheet_name="Long")
                if not df_curves.empty:
                    df_curves.to_excel(writer, index=False, sheet_name="Curves_Polarization")
                # auto-width for all sheets
                for sheet_name, df in [("Wide", df_wide), ("Long", df_long)] + (
                        [("Curves_Polarization", df_curves)] if not df_curves.empty else []):
                    ws = writer.sheets[sheet_name]
                    for col_idx, col_name in enumerate(df.columns, start=1):
                        series = df[col_name].fillna("")
                        max_len = max(len(str(col_name)),
                                      series.astype(str).map(len).max() if not series.empty else 0)
                        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
            messagebox.showinfo("Success", f"Results exported to:\n{save_path}")
            set_status(f"Exported: {os.path.basename(save_path)}")

        elif ext == ".csv":
            wide_path = f"{base}_wide.csv"
            long_path = f"{base}_long.csv"
            df_wide.to_csv(wide_path, index=False, float_format="%.4g")
            df_long.to_csv(long_path, index=False, float_format="%.4g")
            curves_path = None
            if not df_curves.empty:
                curves_path = f"{base}_curves_polarization.csv"
                df_curves.to_csv(curves_path, index=False, float_format="%.4g")
            msg = f"Results exported to:\n{wide_path}\n{long_path}"
            if curves_path:
                msg += f"\n{curves_path}"
            messagebox.showinfo("Success", msg)
            set_status(f"Exported CSV: {os.path.basename(wide_path)}, {os.path.basename(long_path)}"
                       + (f", {os.path.basename(curves_path)}" if curves_path else ""))

        else:
            messagebox.showerror("Error", f"Unsupported extension: {ext}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export results:\n{e}")


# ----------------------------
# Excel Update (with formatting) — match on first 5 chars, round 4 sig figs
# ----------------------------
def update_master_excel():
    """
    Let the user pick a master Excel file, update rows by the first 5 chars of 'MEA'
    using current_at_06V, and save to a new .xlsx with formatting.
    """
    if not current_at_06V:
        messagebox.showinfo("Info", "No data to update. Please click 'Extract Current...' first.")
        return

    if not ensure_openpyxl():
        return

    # 1) Choose Excel file
    file_path = filedialog.askopenfilename(
        title="Select Excel file to update",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file_path:
        return

    # 2) Read
    try:
        df_excel = pd.read_excel(file_path, engine="openpyxl")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read Excel file:\n{e}")
        return

    # 3) Define columns we will write to (pretty 'Pol X' names)
    columns_to_update = []
    for i in range(1, 7):
        columns_to_update.append(f"Pol {i} i @ 0.6 V [A/cm²]")
    for i in range(1, 7):
        columns_to_update.append(f"Pol {i} i @ 0.65 V [A/cm²]")
    for i in range(1, 7):
        columns_to_update.append(f"Pol {i} OCV [V]")
    for i in range(1, 7):
        columns_to_update.append(f"Pol {i} Pmax [W/cm²]")

    # 4) Ensure columns exist — appended at the end if the master file lacks them.
    #    DATA_STRUCTURE_COLUMN is handled apart from columns_to_update everywhere below:
    #    it carries text, and columns_to_update is what the 4-sig-fig rounding runs over,
    #    which would turn every verdict into NaN.
    for col in columns_to_update + [DATA_STRUCTURE_COLUMN]:
        if col not in df_excel.columns:
            df_excel[col] = np.nan

    # 5) Build lookup from results — key on first 5 chars of folder name
    normalized_results = {normalize_str(k)[:5]: v for k, v in current_at_06V.items()}
    # Same 5-char key back to the full folder name, which is what the structure verdict
    # is stored under.
    normalized_keys = {normalize_str(k)[:5]: k for k in current_at_06V}

    # 6) Update loop — match first 5 chars of MEA
    updated_any = False
    for idx, row in df_excel.iterrows():
        mea = normalize_str(row.get("MEA"))[:5]
        if not mea:
            continue
        if mea in normalized_results:
            values = normalized_results[mea]
            for col in columns_to_update:
                if col in values and values[col] != "N/A":
                    df_excel.at[idx, col] = values[col]
                    updated_any = True
            # Written even when no metric matched: a row whose numbers stayed empty is
            # exactly the one where knowing why is worth most.
            df_excel.at[idx, DATA_STRUCTURE_COLUMN] = data_structure_status(normalized_keys[mea])
            updated_any = True

    if not updated_any:
        messagebox.showwarning(
            "No matches",
            "No rows were updated.\n\n"
            "Tips:\n"
            "• Ensure the first 5 characters of 'MEA' in Excel match your folder names (e.g., 'M4813')."
        )
        return

    # Round to 4 sig figs
    for col in columns_to_update:
        if col in df_excel.columns:
            df_excel[col] = df_excel[col].apply(round_sig_numeric)

    # 7) Save with formatting
    from openpyxl.utils import get_column_letter

    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Save updated Excel as..."
    )
    if not save_path:
        return

    try:
        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            sheet_name = "Data"
            df_excel.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            # Freeze top row and first column
            ws.freeze_panes = "B2"  # freezes row 1 and column A

            # Auto-size columns based on header and cell content
            for col_idx, col_name in enumerate(df_excel.columns, start=1):
                series = df_excel[col_name].fillna("")
                max_len = max(
                    len(str(col_name)),
                    series.astype(str).map(len).max() if not series.empty else 0
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        messagebox.showinfo("Success", f"File saved:\n{save_path}")
        set_status(f"Master Excel updated: {os.path.basename(save_path)}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save Excel file:\n{e}")


# ----------------------------
# eLabFTW Export (extra_fields JSON per experiment, grouped by Pol)
# ----------------------------
BOL_CONFIG_FILENAME = "elabftw_bol_procedures.json"
ELABFTW_VALUES_FILENAME = "elabftw_export_values.json"
EXPORT_FORMAT_FILENAME = "elabftw_export_format.json"


def _app_dir() -> str:
    """Folder the user sees: next to the .exe when frozen, next to the .py otherwise.
    Use for files colleagues edit and for anything that must survive a restart -
    a PyInstaller build unpacks itself to a temp dir that is deleted on exit."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _bundle_dir() -> str:
    """Read-only assets shipped inside the build (PyInstaller unpacks to _MEIPASS)."""
    return getattr(sys, "_MEIPASS", _app_dir())


def _find_file(name: str):
    """A copy next to the program wins over the bundled fallback; None if neither."""
    for d in (_app_dir(), _bundle_dir()):
        p = os.path.join(d, name)
        if os.path.exists(p):
            return p
    return None


def load_export_config():
    """
    Read the external, user-editable eLabFTW-format config file
    (elabftw_bol_procedures.json, next to this script) and extract:
      - metadata_fields: field defs from the group "Identification" (once per export run)
      - bol_procedures:  [{"name","fields"}] for groups named "BOL: ..." (once per export run -
                         BOL conditions the whole cell before any Pol curve, not one Pol)
      - ast_procedures:  [{"name","fields"}] for groups named "AST: ..." (once per export run,
                         optional - not every test includes an accelerated stress test)
      - pol_condition_fields: field defs from "EC1 Conditions: *" (stoichiometry/T/p/RH anode
                         & cathode) - the operating conditions of one Pol curve, set per Pol 1..6
    Each field def is {"label", "type", "unit", "options", "default"}.
    Returns ([], [], [], []) if the file is missing or malformed, so the rest of the
    export still works with just the Pol results.
    """
    path = _find_file(BOL_CONFIG_FILENAME)
    if path is None:
        print(f"[load_export_config] {BOL_CONFIG_FILENAME} not found in {_app_dir()}", file=sys.stderr)
        return [], [], [], []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"[load_export_config] Could not read {path}: {e}", file=sys.stderr)
        return [], [], [], []

    extra_fields = data.get("extra_fields", {}) or {}
    groups = data.get("elabftw", {}).get("extra_fields_groups", []) or []
    gid_to_name = {g["id"]: g["name"] for g in groups if "id" in g and "name" in g}
    name_to_gid = {name: gid for gid, name in gid_to_name.items()}

    fields_by_group = {}
    for field_name, spec in extra_fields.items():
        gid = spec.get("group_id")
        fields_by_group.setdefault(gid, []).append((spec.get("position", 0), field_name, spec))
    for gid in fields_by_group:
        fields_by_group[gid].sort(key=lambda t: t[0])

    def _field_def(field_name, spec, sub_label=None):
        return {
            # "label" is the short form shown in the export dialog ("Temperature"); "name" is
            # the procedure file's own field name ("BOL: Galvanostatic Ramp (GSTEP):
            # Temperature"). Keeping both is what lets the export write the name the
            # Procedure Builder already used, instead of inventing a parallel one that then
            # has to be mapped back. See build_elabftw_payload/add_fields.
            "label": sub_label if sub_label is not None else field_name,
            "name": field_name,
            "type": spec.get("type", "text"),
            "unit": spec.get("unit", ""),
            "options": spec.get("options"),
            "default": spec.get("value", ""),
        }

    # Two Identification fields are the Procedure Builder's own bookkeeping, not things an
    # operator should be asked to retype: "Condition Catalog Size" defines how that tool
    # numbers post-AST blocks, and "Export Format" records which naming format wrote the
    # procedure file. Both belong to the procedure and are already carried by it; offering
    # them here would invite a value that contradicts the file they came from.
    _SKIP_METADATA = {"Condition Catalog Size", "Export Format"}
    metadata_fields = [
        _field_def(fname, spec)
        for _, fname, spec in fields_by_group.get(name_to_gid.get("Identification"), [])
        if fname not in _SKIP_METADATA
    ]

    def _procedures_by_prefix(prefix, skip_labels):
        out = []
        for g in groups:
            gname = g.get("name", "")
            if not gname.startswith(prefix):
                continue
            sub_fields = []
            for _, fname, spec in fields_by_group.get(g["id"], []):
                sub_label = fname.split(": ")[-1]
                if sub_label in skip_labels:
                    continue  # redundant with the dropdown choice itself
                sub_fields.append(_field_def(fname, spec, sub_label=sub_label))
            out.append({"name": gname, "fields": sub_fields})
        return out

    bol_procedures = _procedures_by_prefix("BOL:", {"Activation Type"})
    ast_procedures = _procedures_by_prefix("AST:", {"AST Type"})

    # Condition fields are now named "EC1 <parameter>" with no "Conditions:" scope word, so
    # a name prefix alone can no longer tell them from "EC1 OCV [V]" or "EC1 CV 1.1 ECSA".
    # They are found by their GROUP instead - the builder emits "Test Conditions - EC1" in
    # both compact and full export - and the sub-label is what remains after stripping the
    # "EC1 " block prefix.
    #
    # Skipped: "Condition ID" is the builder's bookkeeping - which catalog entry the block is
    # - which this tool already encodes in the Pol number it exports as EC{n}. "Same As"
    # describes a post-AST repeat and is decided per block, not typed once for all six.
    _skip_conditions = {"Condition ID", "Same As"}
    _ec1_conditions_gid = next(
        (gid for gid, name in gid_to_name.items()
         if name.replace("—", "-").strip().startswith("Test Conditions")
         and name.rstrip().endswith("EC1")),
        None,
    )
    pol_condition_fields = [
        _field_def(fname, spec, sub_label=fname[len("EC1 "):])
        for _, fname, spec in fields_by_group.get(_ec1_conditions_gid, [])
        if fname.startswith("EC1 ") and fname[len("EC1 "):] not in _skip_conditions
    ]
    if _ec1_conditions_gid is None:
        print("[load_export_config] No 'Test Conditions - EC1' group found - per-Pol "
              "condition fields will be empty. Re-export elabftw_bol_procedures.json "
              "from the Procedure Builder.", file=sys.stderr)

    return metadata_fields, bol_procedures, ast_procedures, pol_condition_fields


# Shape of the exported eLabFTW JSON: group ids/names, field labels and prefixes.
# Distinct from elabftw_bol_procedures.json (loaded above), which controls what fields
# the export DIALOG offers - this controls what the WRITTEN JSON looks like.
# Kept as the single source of truth for every hardcoded id/label/prefix that used to
# live inline in build_elabftw_payload(), so a missing or partial
# elabftw_export_format.json still reproduces exactly this output.
DEFAULT_EXPORT_FORMAT = {
    "mea_field_name": "MEA",
    "identification_group_id": 9,
    "identification_group_name": "Identification",
    "bol_group_id": 7,
    "ast_group_id": 8,
    # Empty prefixes mean "write the field name the procedure file already used", e.g.
    # "BOL: Galvanostatic Ramp (GSTEP): Temperature". Set one back to "BOL " to return to
    # the old self-invented naming - but then results no longer match the procedure they
    # belong to, which is the whole reason the prefixes were emptied.
    "bol_field_prefix": "",
    "ast_field_prefix": "",
    # Pol N's group id = base + N. Results default to 1..6 (base 0); conditions to
    # 21..26 (base 20) - kept apart so the two never collide.
    "pol_results_group_id_base": 0,
    "pol_conditions_group_id_base": 20,
    # The exported name for Pol N. The Procedure Builder numbers its blocks
    # EC(conditionId + catalogSize × phaseIndex), which is the same arithmetic the bench's
    # SetMarker uses and therefore the same number as this tool's Pol slot - so Pol 4 is
    # EC4, and the two files line up field for field with no mapping table.
    "pol_export_label_pattern": "EC{pol}",
    "pol_results_group_name_pattern": "Key Results — {pol_export_label}",
    "pol_conditions_group_name_pattern": "Test Conditions — {pol_export_label}",
    # No "Conditions: " scope word - the group heading already says it, and repeating it
    # made every row 12 characters longer than the value it carried.
    "pol_conditions_field_prefix_pattern": "{pol_export_label} ",
    # Exported label per result metric - what appears after "EC4 " in the field name. Must
    # match the placeholders the Procedure Builder writes into its Results group, or the
    # measured value lands in a new field instead of the one waiting for it.
    # The metric ids (i06/i065/ocv/pmax) are internal and not meant to be renamed.
    # Units appear in the name as well as the unit property below: flat reports and CSV
    # exports collapse {"value","unit"} to the bare number, and a column of unlabelled
    # magnitudes is unreadable.
    "result_metric_labels": {
        "i06": "i @ 0.6 V [A/cm²]",
        "i065": "i @ 0.65 V [A/cm²]",
        "ocv": "OCV [V]",
        "pmax": "Pmax [W/cm²]",
    },
    # Units belong in eLabFTW's own unit/units properties, not baked into the field name
    # ("Pmax [W/cm²]") where nothing can read them back.
    "result_metric_units": {
        "i06": "A/cm²",
        "i065": "A/cm²",
        "ocv": "V",
        "pmax": "W/cm²",
    },
    "display_main_text": False,
    "filename_pattern": "{key}_eLabFTW.json",
}


def load_export_format() -> dict:
    """External, user-editable export-SHAPE config (elabftw_export_format.json, next to
    the program). Missing keys - or a missing file entirely - fall back to
    DEFAULT_EXPORT_FORMAT one at a time, so a single value can be overridden without
    repeating the rest of the file.
    """
    nested_keys = ("result_metric_labels", "result_metric_units")

    def _fresh_defaults():
        base = dict(DEFAULT_EXPORT_FORMAT)
        for k in nested_keys:
            base[k] = dict(DEFAULT_EXPORT_FORMAT[k])
        return base

    fmt = _fresh_defaults()

    path = _find_file(EXPORT_FORMAT_FILENAME)
    if path is not None:
        try:
            with open(path, "r", encoding="utf-8") as f:
                overrides = json.load(f)
            fmt.update({k: v for k, v in overrides.items() if k not in nested_keys})
            for k in nested_keys:
                fmt[k].update(overrides.get(k, {}) or {})
        except Exception as e:
            print(f"[load_export_format] Could not read {path}, using defaults: {e}", file=sys.stderr)
            fmt = _fresh_defaults()

    # eLabFTW's UI treats a group_id of 0 as falsy ("no group") and dumps the field into
    # an "undefined" bucket - guard the three fixed ids against a stray edit. (The two
    # Pol id *bases* are documented in the shipped file instead: keeping Pol 1..6 away
    # from 0 depends on the base, not a single fixed value.)
    for gid_key in ("identification_group_id", "bol_group_id", "ast_group_id"):
        if not fmt.get(gid_key):
            print(f"[load_export_format] '{gid_key}' must not be 0 - using default", file=sys.stderr)
            fmt[gid_key] = DEFAULT_EXPORT_FORMAT[gid_key]

    # A pattern silently missing its placeholder still "works" (str.format ignores an
    # unused kwarg), just wrong: every batch-exported file would collide on one name and
    # overwrite the last, or every Pol's condition group would share one name.
    if "{key}" not in str(fmt.get("filename_pattern", "")):
        print("[load_export_format] 'filename_pattern' must contain '{key}' - using default", file=sys.stderr)
        fmt["filename_pattern"] = DEFAULT_EXPORT_FORMAT["filename_pattern"]
    if "{pol}" not in str(fmt.get("pol_export_label_pattern", "")):
        print("[load_export_format] 'pol_export_label_pattern' must contain '{pol}' - using default", file=sys.stderr)
        fmt["pol_export_label_pattern"] = DEFAULT_EXPORT_FORMAT["pol_export_label_pattern"]
    for pattern_key in ("pol_results_group_name_pattern", "pol_conditions_group_name_pattern",
                        "pol_conditions_field_prefix_pattern"):
        if "{pol_export_label}" not in str(fmt.get(pattern_key, "")):
            print(f"[load_export_format] '{pattern_key}' must contain '{{pol_export_label}}' - using default", file=sys.stderr)
            fmt[pattern_key] = DEFAULT_EXPORT_FORMAT[pattern_key]

    return fmt


def load_elabftw_saved_values() -> dict:
    """Last-used metadata/BOL values, remembered between export runs."""
    try:
        with open(os.path.join(_app_dir(), ELABFTW_VALUES_FILENAME), "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_elabftw_saved_values(values: dict):
    try:
        with open(os.path.join(_app_dir(), ELABFTW_VALUES_FILENAME), "w", encoding="utf-8") as f:
            json.dump(values, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"[save_elabftw_saved_values] Failed to save: {e}", file=sys.stderr)


_DATE_INPUT_FORMATS = ("%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y")


def _normalize_value(fv: dict) -> str:
    """eLabFTW stores a "date" field as ISO yyyy-mm-dd and shows anything else as an empty
    date picker, so the German dd.mm.yyyy the dialog is usually filled with would silently
    arrive blank. Non-date fields, and dates already ISO or in an unrecognised shape, pass
    through untouched rather than being guessed at."""
    value = fv.get("value", "")
    if fv.get("type") != "date" or not isinstance(value, str) or not value.strip():
        return value
    text = value.strip()
    try:
        datetime.strptime(text, "%Y-%m-%d")
        return text
    except ValueError:
        pass
    for fmt_str in _DATE_INPUT_FORMATS:
        try:
            return datetime.strptime(text, fmt_str).strftime("%Y-%m-%d")
        except ValueError:
            continue
    print(f"[_normalize_value] Unrecognised date '{text}' - exporting as-is", file=sys.stderr)
    return text


def build_elabftw_payload(key: str, metadata_values: dict = None, bol_values: dict = None,
                           ast_values: dict = None, pol_conditions: dict = None,
                           export_format: dict = None) -> dict:
    """Build an eLabFTW-compatible {"extra_fields": ..., "elabftw": {...}} dict for one
    experiment/key. Each Pol with data becomes its own extra_fields_group.

    metadata_values: {label: {"value","type","unit","options"}} - once per export run.
    bol_values / ast_values: {"procedure": <name>, "params": {label: {"value","type","unit","options"}}}
                              - once per export run (BOL/AST condition the whole cell, not one Pol).
    pol_conditions: {pol_number: {label: {"value","type","unit","options"}}}
                     - per-Pol operating conditions (stoichiometry, T, p, RH anode/cathode).
    export_format: shape of the OUTPUT - group ids/names, labels, prefixes (see
                    load_export_format()). Loaded fresh if not given.
    """
    fmt = export_format if export_format is not None else load_export_format()
    values = current_at_06V.get(key, {})

    def first_present(d, keys):
        for k in keys:
            if k in d:
                return d[k]
        return "N/A"

    def clean(val):
        if val == "N/A" or val is None or (isinstance(val, float) and np.isnan(val)):
            return None
        rounded = round_sig_numeric(val)
        return None if pd.isna(rounded) else rounded

    def result_lookup_keys(pol_lbl: str) -> dict:
        """Internal lookup keys into current_at_06V, per metric id. Fixed - must match
        exactly what extract_current() writes there, so this stays independent of the
        (configurable) exported label below."""
        pol_compact = pol_lbl.replace(" ", "")
        return {
            "i06": [f"{pol_lbl} i @ 0.6 V [A/cm²]", f"{pol_compact} i @ 0.6 V [A/cm²]", f"{pol_compact} Current @ 0.6V [A]"],
            "i065": [f"{pol_lbl} i @ 0.65 V [A/cm²]", f"{pol_compact} i @ 0.65 V [A/cm²]"],
            "ocv": [f"{pol_lbl} OCV [V]", f"{pol_compact} OCV [V]"],
            "pmax": [f"{pol_lbl} Pmax [W/cm²]", f"{pol_compact} Pmax [W/cm²]"],
        }

    # NB: group_id must never be 0 - eLabFTW's UI treats 0 as falsy/"no group" and
    # dumps the field into an "undefined" bucket, same as if group_id were missing.
    identification_gid = fmt["identification_group_id"]
    extra_fields = {fmt["mea_field_name"]: {"type": "text", "value": key, "position": 1, "group_id": identification_gid}}
    groups = [{"id": identification_gid, "name": fmt["identification_group_name"]}]
    position = 2

    def add_fields(field_values, group_id, group_name, name_prefix=""):
        nonlocal position
        if not field_values:
            return
        if not any(g["id"] == group_id for g in groups):
            groups.append({"id": group_id, "name": group_name})
        for label, fv in field_values.items():
            entry = {"type": fv.get("type", "text"), "value": _normalize_value(fv), "position": position, "group_id": group_id}
            if fv.get("unit"):
                entry["unit"] = fv["unit"]
                entry["units"] = [fv["unit"]]
            if fv.get("options"):
                entry["options"] = fv["options"]
            # A prefix means the field is one of a repeated set (Pol 1..6 conditions), where
            # the procedure file only supplied the EC1 spelling - re-prefix per Pol. Without
            # one the field is a singleton (metadata, BOL, AST) and the procedure file's own
            # name is used verbatim, so both tools write the identical key.
            if name_prefix:
                field_name = f"{name_prefix}{label}"
            else:
                field_name = fv.get("name") or label
            extra_fields[field_name] = entry
            position += 1

    # ---- Metadata (Identification) - once per export run ----
    add_fields(metadata_values, identification_gid, fmt["identification_group_name"])

    # ---- BOL / AST procedure - once per export run, conditions the whole cell ----
    if bol_values and bol_values.get("params"):
        add_fields(bol_values["params"], fmt["bol_group_id"], bol_values["procedure"], name_prefix=fmt["bol_field_prefix"])
    if ast_values and ast_values.get("params"):
        add_fields(ast_values["params"], fmt["ast_group_id"], ast_values["procedure"], name_prefix=fmt["ast_field_prefix"])

    # ---- Per-Pol results + operating conditions ----
    for pol in POL_CODES:
        # Two different labels on purpose: pol_lbl ("Pol 3") is the internal key
        # extract_current() wrote into current_at_06V and must not change, while export_lbl
        # ("EC3") is what goes into the file, matching the Procedure Builder's block naming.
        pol_lbl = f"Pol {pol}"
        export_lbl = fmt["pol_export_label_pattern"].format(pol=pol)

        pol_fields = {}
        for metric_id, lookup_keys in result_lookup_keys(pol_lbl).items():
            val = clean(first_present(values, lookup_keys))
            if val is None:
                continue
            exported_label = fmt["result_metric_labels"].get(
                metric_id, DEFAULT_EXPORT_FORMAT["result_metric_labels"][metric_id])
            unit = fmt["result_metric_units"].get(
                metric_id, DEFAULT_EXPORT_FORMAT["result_metric_units"].get(metric_id, ""))
            pol_fields[f"{export_lbl} {exported_label}"] = (val, unit)

        if pol_fields:
            pol_gid = fmt["pol_results_group_id_base"] + pol
            groups.append({"id": pol_gid,
                           "name": fmt["pol_results_group_name_pattern"].format(pol_export_label=export_lbl)})
            for field_name, (val, unit) in pol_fields.items():
                entry = {
                    "type": "number",
                    "value": str(val),
                    "position": position,
                    "group_id": pol_gid,
                }
                if unit:
                    entry["unit"] = unit
                    entry["units"] = [unit]
                extra_fields[field_name] = entry
                position += 1

            # Only export this Pol's operating conditions if it actually has a curve
            cond_gid = fmt["pol_conditions_group_id_base"] + pol
            add_fields(
                (pol_conditions or {}).get(pol),
                cond_gid,
                fmt["pol_conditions_group_name_pattern"].format(pol_export_label=export_lbl),
                name_prefix=fmt["pol_conditions_field_prefix_pattern"].format(pol_export_label=export_lbl),
            )

    return {
        "extra_fields": extra_fields,
        "elabftw": {
            "display_main_text": fmt["display_main_text"],
            "extra_fields_groups": groups,
        },
    }


def export_to_elabftw(keys, metadata_values: dict = None, bol_values: dict = None,
                       ast_values: dict = None, pol_conditions: dict = None):
    """Write one eLabFTW-compatible JSON file per key in `keys`."""
    if not current_at_06V:
        messagebox.showinfo("Info", "No results to export. Run 'Extract i@0.6 / i@0.65 + OCV + Pmax' first.")
        return

    keys = [k for k in keys if k in current_at_06V]
    if not keys:
        messagebox.showwarning("Export to eLabFTW", "No matching datasets to export.")
        return

    # Load once per export run, not once per key - and share it so every key in this
    # run gets the same shape, and a Pol's result group id can be told apart from any
    # other group below without assuming the old fixed 1..6 range.
    fmt = load_export_format()
    pol_result_gids = {fmt["pol_results_group_id_base"] + p for p in POL_CODES}

    if len(keys) == 1:
        key = keys[0]
        save_path = filedialog.asksaveasfilename(
            title="Save eLabFTW JSON as...",
            defaultextension=".json",
            initialfile=fmt["filename_pattern"].format(key=key),
            filetypes=[("JSON", "*.json")],
        )
        if not save_path:
            return
        try:
            payload = build_elabftw_payload(key, metadata_values, bol_values, ast_values, pol_conditions, export_format=fmt)
            with open(save_path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, ensure_ascii=False)
            messagebox.showinfo("Success", f"eLabFTW JSON exported to:\n{save_path}")
            set_status(f"Exported eLabFTW JSON: {os.path.basename(save_path)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export eLabFTW JSON:\n{e}")
        return

    outdir = filedialog.askdirectory(title="Choose output folder for eLabFTW JSON files")
    if not outdir:
        return

    written, empty = [], []
    for key in keys:
        payload = build_elabftw_payload(key, metadata_values, bol_values, ast_values, pol_conditions, export_format=fmt)
        safe_name = re.sub(r'[\\/:*?"<>|]', "_", key)
        path = os.path.join(outdir, fmt["filename_pattern"].format(key=safe_name))
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(payload, f, indent=2, ensure_ascii=False)
            written.append(path)
            has_pol_results = any(g["id"] in pol_result_gids for g in payload["elabftw"]["extra_fields_groups"])
            if not has_pol_results:
                empty.append(key)
        except Exception as e:
            print(f"[export_to_elabftw] Failed to write {path}: {e}", file=sys.stderr)

    msg = f"Wrote {len(written)} eLabFTW JSON file(s) to:\n{outdir}"
    if empty:
        msg += f"\n\nNo Pol data found for: {', '.join(empty)}"
    messagebox.showinfo("Success", msg)
    set_status(f"Exported {len(written)} eLabFTW JSON file(s) to '{os.path.basename(outdir)}'")


def _make_scrollable(parent, height=420, width=780):
    """Canvas+Scrollbar based scrollable frame. Returns (container_to_pack, inner_frame_to_fill)."""
    container = tk.Frame(parent)
    canvas = tk.Canvas(container, height=height, width=width, highlightthickness=0)
    vbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
    inner = tk.Frame(canvas)

    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner, anchor="nw")
    canvas.configure(yscrollcommand=vbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    vbar.pack(side="right", fill="y")
    return container, inner


def _add_field_row(parent, row, col, fdef, var):
    """Grid a Label+input (Combobox for 'select' types, Entry otherwise) for one field def."""
    tk.Label(parent, text=f"{fdef['label']}:").grid(row=row, column=col, sticky="e", padx=4, pady=2)
    if fdef.get("type") == "select" and fdef.get("options"):
        w = ttk.Combobox(parent, textvariable=var, values=fdef["options"], state="readonly", width=18)
    else:
        w = tk.Entry(parent, textvariable=var, width=20)
    w.grid(row=row, column=col + 1, sticky="w", padx=4, pady=2)


def _build_procedure_box(parent, title, procedures, saved_choice):
    """One dataset-wide dropdown-driven procedure box (used for BOL and AST alike, since both
    condition the whole cell once, not one Pol curve). Returns a dict with the live Tk vars /
    field defs the caller reads back after the Export button is pressed."""
    names = [p["name"] for p in procedures]
    by_name = {p["name"]: p for p in procedures}

    box = tk.LabelFrame(parent, text=title)
    box.pack(fill="x", pady=(0, 10), padx=4)

    top_row = tk.Frame(box)
    top_row.pack(fill="x", padx=6, pady=(4, 2))
    tk.Label(top_row, text="Procedure:").pack(side="left")
    proc_var = tk.StringVar(value=saved_choice.get("procedure", ""))
    combo = ttk.Combobox(top_row, textvariable=proc_var, values=[""] + names, state="readonly", width=42)
    combo.pack(side="left", padx=6)

    fields_frame = tk.Frame(box)
    fields_frame.pack(fill="x", padx=6, pady=(2, 6))

    state = {"proc_var": proc_var, "param_vars": {}, "field_defs": {}}

    def render_fields(prefill=None):
        for child in fields_frame.winfo_children():
            child.destroy()
        state["param_vars"] = {}
        state["field_defs"] = {}
        proc = by_name.get(proc_var.get())
        if not proc:
            return
        prefill = prefill or {}
        for i, fdef in enumerate(proc["fields"]):
            label = fdef["label"]
            default = prefill.get(label, fdef.get("default", ""))
            var = tk.StringVar(value=str(default))
            state["param_vars"][label] = var
            state["field_defs"][label] = fdef
            r, c = divmod(i, 2)
            _add_field_row(fields_frame, r, c * 2, fdef, var)

    combo.bind("<<ComboboxSelected>>", lambda e: render_fields())
    if proc_var.get() in by_name:
        render_fields(prefill=saved_choice.get("params", {}))

    return state


def open_elabftw_export_dialog():
    """Dialog: export scope (selected/all) + Identification metadata + BOL/AST procedure
    (once per dataset - they condition the whole cell before any Pol curve) + per-Pol operating
    conditions (stoichiometry, T, p, RH anode/cathode). All field definitions are sourced from
    elabftw_bol_procedures.json, so new fields/procedures need no code changes. Values are
    remembered between runs via elabftw_export_values.json."""
    if not current_at_06V:
        messagebox.showinfo("Info", "No results to export. Run 'Extract i@0.6 / i@0.65 + OCV + Pmax' first.")
        return

    metadata_fields, bol_procedures, ast_procedures, pol_condition_fields = load_export_config()
    saved = load_elabftw_saved_values()

    win = tk.Toplevel(root)
    win.title("Export to eLabFTW (JSON)")
    win.grab_set()

    # ---- Scope: selected dataset vs. all ----
    selected_key = key_dropdown.get()
    scope_var = tk.StringVar(value="selected" if selected_key in current_at_06V else "all")
    scope_frame = tk.Frame(win)
    scope_frame.pack(fill="x", padx=12, pady=(12, 4))
    sel_label = f"Selected dataset ({selected_key})" if selected_key in current_at_06V else "Selected dataset (none selected)"
    tk.Radiobutton(scope_frame, text=sel_label, variable=scope_var, value="selected",
                   state=("normal" if selected_key in current_at_06V else "disabled")).pack(anchor="w")
    tk.Radiobutton(scope_frame, text=f"All datasets ({len(current_at_06V)})", variable=scope_var, value="all").pack(anchor="w")

    container, inner = _make_scrollable(win)
    container.pack(fill="both", expand=True, padx=12, pady=4)

    if not any([metadata_fields, bol_procedures, ast_procedures, pol_condition_fields]):
        tk.Label(
            inner,
            text=f"({BOL_CONFIG_FILENAME} not found next to the program – exporting Pol results only)",
            fg="#b00000",
        ).pack(anchor="w", pady=6)

    # ---- Metadata (Identification) - once per export run ----
    meta_vars = {}
    if metadata_fields:
        saved_meta = saved.get("metadata", {})
        meta_box = tk.LabelFrame(inner, text="Metadata")
        meta_box.pack(fill="x", pady=(0, 10), padx=4)
        for i, fdef in enumerate(metadata_fields):
            label = fdef["label"]
            default = saved_meta.get(label, fdef.get("default", ""))
            if label == "Active Area" and not saved_meta.get(label):
                default = active_area_var.get()  # reuse the app's own active-area setting
            var = tk.StringVar(value=str(default))
            meta_vars[label] = var
            _add_field_row(meta_box, i, 0, fdef, var)

    # ---- BOL / AST procedure - once per export run, conditions the whole cell ----
    bol_state = _build_procedure_box(inner, "BOL procedure (once per dataset)", bol_procedures,
                                      saved.get("bol", {})) if bol_procedures else None
    ast_state = _build_procedure_box(inner, "AST procedure (once per dataset, optional)", ast_procedures,
                                      saved.get("ast", {})) if ast_procedures else None

    # ---- Per-Pol operating conditions (stoichiometry, T, p, RH anode/cathode) ----
    pol_cond_widgets = {}
    if pol_condition_fields:
        saved_conditions = saved.get("pol_conditions", {})
        for pol in range(1, 7):
            saved_c = saved_conditions.get(str(pol), {})
            box = tk.LabelFrame(inner, text=f"Pol {pol} conditions")
            box.pack(fill="x", pady=(0, 10), padx=4)
            vars_ = {}
            for i, fdef in enumerate(pol_condition_fields):
                label = fdef["label"]
                default = saved_c.get(label, fdef.get("default", ""))
                var = tk.StringVar(value=str(default))
                vars_[label] = var
                r, c = divmod(i, 3)
                _add_field_row(box, r, c * 2, fdef, var)
            pol_cond_widgets[pol] = vars_

    # ---- Buttons ----
    btns = tk.Frame(win)
    btns.pack(pady=10)

    def _collect(field_defs_by_label, vars_by_label):
        out = {}
        for label, var in vars_by_label.items():
            v = var.get().strip()
            if not v:
                continue
            fdef = field_defs_by_label[label]
            out[label] = {"value": v, "type": fdef.get("type", "text"), "unit": fdef.get("unit", ""),
                          "options": fdef.get("options"), "name": fdef.get("name")}
        return out

    def do_export():
        if scope_var.get() == "selected":
            keys = [selected_key]
        else:
            keys = list(current_at_06V.keys())

        metadata_fdefs = {f["label"]: f for f in metadata_fields}
        metadata_values = _collect(metadata_fdefs, meta_vars)

        bol_values = None
        if bol_state and bol_state["proc_var"].get():
            bol_values = {"procedure": bol_state["proc_var"].get(),
                          "params": _collect(bol_state["field_defs"], bol_state["param_vars"])}

        ast_values = None
        if ast_state and ast_state["proc_var"].get():
            ast_values = {"procedure": ast_state["proc_var"].get(),
                          "params": _collect(ast_state["field_defs"], ast_state["param_vars"])}

        cond_fdefs = {f["label"]: f for f in pol_condition_fields}
        pol_conditions = {}
        for pol, vars_ in pol_cond_widgets.items():
            collected = _collect(cond_fdefs, vars_)
            if collected:
                pol_conditions[pol] = collected

        save_elabftw_saved_values({
            "metadata": {k: v["value"] for k, v in metadata_values.items()},
            "bol": ({"procedure": bol_values["procedure"], "params": {k: v["value"] for k, v in bol_values["params"].items()}}
                    if bol_values else {}),
            "ast": ({"procedure": ast_values["procedure"], "params": {k: v["value"] for k, v in ast_values["params"].items()}}
                    if ast_values else {}),
            "pol_conditions": {str(p): {k: v["value"] for k, v in c.items()} for p, c in pol_conditions.items()},
        })

        win.destroy()
        export_to_elabftw(keys, metadata_values, bol_values, ast_values, pol_conditions)

    tk.Button(btns, text="Export…", command=do_export).pack(side="left", padx=6)
    tk.Button(btns, text="Cancel", command=win.destroy).pack(side="left", padx=6)


# ----------------------------
# GUI Actions
# ----------------------------
def _populate_from(root_path: str, source_label: str):
    """Scrape a root folder and fill the dropdown with the keys found."""
    global data_dict
    data_dict = scrape_data(root_path)
    key_dropdown['values'] = list(data_dict.keys())

    if not data_dict:
        messagebox.showwarning(
            "No data found",
            f"No readable measurement files (…01_YYYYMMDD.txt, or 02_/03_ as fallback) "
            f"were found in '{source_label}'."
        )
        set_status(f"No data found in '{source_label}'")
        return

    messagebox.showinfo("Info", f"Data scraped!\nKeys: {list(data_dict.keys())}")
    set_status(f"Loaded {len(data_dict)} keys from '{source_label}'")


def load_from_folder():
    """Pick a root folder of measurement data and populate the dropdown."""
    folder_selected = filedialog.askdirectory(title="Select data folder")
    if not folder_selected:
        return
    _populate_from(folder_selected, os.path.basename(folder_selected))


def load_from_archive():
    """Pick a .rar/.zip archive, unpack it to a temp folder, read it, and clean up."""
    archive = filedialog.askopenfilename(
        title="Select a .rar / .zip archive",
        filetypes=[("Data archives", "*.rar *.zip"), ("All files", "*.*")],
    )
    if not archive:
        return

    name = os.path.basename(archive)
    set_status(f"Extracting '{name}' …")
    root.update_idletasks()          # repaint before the (possibly slow) unpack
    tmp_root = tempfile.mkdtemp(prefix="yalla_")
    try:
        extracted = extract_archive(archive, tmp_root)
        # Safe to delete the temp copy right after: scrape_data() reads every file
        # into pandas and keeps no handles.
        _populate_from(extracted, name)
    except RuntimeError as e:
        messagebox.showerror("Archive", str(e))
        set_status("Archive could not be opened")
    finally:
        shutil.rmtree(tmp_root, ignore_errors=True)

def on_key_select():
    """Plot Polarization curves for selected key."""
    selected_key = key_dropdown.get()
    if selected_key in data_dict:
        plot_data(selected_key, data_dict[selected_key])

def on_power_plot():
    """Plot power curves for selected key."""
    selected_key = key_dropdown.get()
    if selected_key in data_dict:
        plot_power_data(selected_key, data_dict[selected_key])


# ----------------------------
# Bar Diagram (Extracted Values)
# ----------------------------
def plot_bar_diagram_multi_pol(keys, metric, pols, metric_label, key_labels=None):
    """Plot bar diagram for a single metric across multiple keys and Pols.
    
    Each key group shows bars for each selected Pol.
    """
    if not keys or not metric or not pols:
        messagebox.showwarning("Bar Diagram", "Missing keys, metric, or Pols.")
        return
    
    # Prepare data: for each key, collect values for each Pol
    data_by_pol = {pol: [] for pol in pols}
    labels = []
    
    # Track missing data combinations
    missing_data = []
    
    for key in keys:
        if key not in current_at_06V:
            continue
        
        values = current_at_06V[key]
        labels.append(key)
        
        for pol in pols:
            pol_lbl = f"Pol {pol}"
            col_name = f"{pol_lbl} {metric}"
            val = values.get(col_name, "N/A")
            
            # Convert to float or NaN
            if val == "N/A" or val is None or (isinstance(val, float) and np.isnan(val)):
                data_by_pol[pol].append(np.nan)
                missing_data.append((key, pol))
            else:
                try:
                    data_by_pol[pol].append(float(val))
                except Exception:
                    data_by_pol[pol].append(np.nan)
                    missing_data.append((key, pol))
    
    if not labels:
        messagebox.showwarning("Bar Diagram", "No data found for selected keys and Pols.")
        return
    
    # Warn about missing data
    if missing_data:
        missing_str = ", ".join([f"Key '{k}' Pol {p}" for k, p in missing_data])
        messagebox.showwarning("Bar Diagram", f"No data for: {missing_str}")
    
    # Determine x-axis labels
    if key_labels and len(key_labels) == len(labels):
        x_labels = key_labels
    else:
        x_labels = labels
    
    # Create bar diagram
    fig, ax = plt.subplots(figsize=(12, 6))
    
    num_pols = len(pols)
    bar_width = 0.15
    group_gap = 0.1  # small gap between key groups
    x = np.arange(len(labels)) * (num_pols * bar_width + group_gap)
    
    # Colors for different Pols
    colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b"]
    
    for idx, pol in enumerate(sorted(pols)):
        offset = (idx - (num_pols - 1) / 2) * bar_width
        ax.bar(x + offset, data_by_pol[pol], bar_width, label=f"Pol {pol}", color=colors[idx % len(colors)])
    
    # Get tick label fontsize
    tick_fontsize = ax.xaxis.get_ticklabels()[0].get_fontsize() if ax.xaxis.get_ticklabels() else 12
    
    ax.set_xlabel("", fontsize=tick_fontsize)  # no default x-label
    ax.set_ylabel(metric_label, fontsize=tick_fontsize)
    ax.set_title(f"{metric_label}", fontsize=tick_fontsize, fontweight="bold")
    ax.set_xticks(x)
    ax.set_xticklabels(x_labels, rotation=45, ha="right", fontsize=tick_fontsize)
    
    leg = ax.legend(loc="upper left", frameon=True)
    leg.set_draggable(True)
    leg.get_frame().set_alpha(0.9)
    leg.get_frame().set_facecolor("white")
    leg.get_frame().set_edgecolor("black")
    for text in leg.get_texts():
        text.set_fontsize(tick_fontsize)
    
    ax.grid(True, alpha=0.3, axis="y")
    
    plt.tight_layout()
    set_status(f"Bar Diagram | {metric_label} | keys={len(labels)} | pols={len(pols)}")
    attach_label_editor(fig, ax)
    plt.show()


def open_bar_selector():
    """Dialog to select keys and metrics for bar diagram plotting."""
    if not current_at_06V:
        messagebox.showwarning("Bar Diagram", "No extracted data available. Run 'Extract Current...' first.")
        return
    
    win = tk.Toplevel(root)
    win.title("Bar Diagram: choose Keys, Metric & Pols")
    win.grab_set()
    
    # Left: Keys checkboxes + rename fields
    tk.Label(win, text="Keys").grid(row=0, column=0, padx=8, pady=(8, 2), sticky="w")
    tk.Label(win, text="Rename Keys").grid(row=0, column=1, padx=8, pady=(8, 2), sticky="w")
    
    keys_list = sorted(current_at_06V.keys())
    key_vars = {}
    key_label_vars = {}
    
    for idx, key in enumerate(keys_list):
        var = tk.BooleanVar(value=(idx == 0))  # default: select first key
        key_vars[key] = var
        row = 1 + idx
        
        tk.Checkbutton(win, text=key, variable=var).grid(row=row, column=0, sticky="w", padx=8, pady=1)
        
        lbl_var = tk.StringVar(value=key)
        key_label_vars[key] = lbl_var
        tk.Entry(win, textvariable=lbl_var, width=16).grid(row=row, column=1, sticky="w", padx=4, pady=1)
    
    # Middle: Pol checkboxes
    tk.Label(win, text="Pols").grid(row=0, column=2, padx=8, pady=(8, 2), sticky="w")
    
    pol_vars = {}
    for idx, pol in enumerate(POL_CODES):
        var = tk.BooleanVar(value=(pol == 1))  # default: select Pol 1
        pol_vars[pol] = var
        tk.Checkbutton(win, text=f"Pol {pol}", variable=var).grid(row=1+idx, column=2, sticky="w", padx=8, pady=2)
    
    # Right: Metric dropdown
    tk.Label(win, text="Metric").grid(row=0, column=3, padx=8, pady=(8, 2), sticky="w")
    
    metrics_available = [
        "i @ 0.6 V [A/cm²]",
        "i @ 0.65 V [A/cm²]",
        "OCV [V]",
        "Pmax [W/cm²]"
    ]
    
    metric_choice = tk.StringVar(value=metrics_available[0])
    metric_label_var = tk.StringVar(value=metrics_available[0])
    
    def on_metric_change(*args):
        """Update the rename field when metric selection changes."""
        metric_label_var.set(metric_choice.get())
    
    cb_metric = ttk.Combobox(win, textvariable=metric_choice, values=metrics_available,
                             state="readonly", width=18)
    cb_metric.grid(row=1, column=3, padx=8, pady=2, sticky="nw")
    metric_choice.trace("w", on_metric_change)
    
    # Metric rename field
    tk.Label(win, text="Rename:").grid(row=2, column=3, padx=8, pady=(8, 2), sticky="w")
    tk.Entry(win, textvariable=metric_label_var, width=18).grid(row=3, column=3, sticky="nw", padx=8, pady=2)
    
    # Bottom buttons
    btns = tk.Frame(win)
    btns.grid(row=max(1 + len(keys_list), 8), column=0, columnspan=4, pady=20)
    
    def collect_selection():
        sel_keys = [k for k, v in key_vars.items() if v.get()]
        sel_pols = [p for p, v in pol_vars.items() if v.get()]
        metric = metric_choice.get()
        metric_label = metric_label_var.get()
        sel_key_labels = [key_label_vars[k].get() for k in sel_keys]
        return sel_keys, sel_pols, metric, metric_label, sel_key_labels
    
    def plot_now():
        sel_keys, sel_pols, metric, metric_label, sel_key_labels = collect_selection()
        if not sel_keys:
            messagebox.showwarning("Bar Diagram", "Select at least one key.")
            return
        if not sel_pols:
            messagebox.showwarning("Bar Diagram", "Select at least one Pol.")
            return
        plot_bar_diagram_multi_pol(sel_keys, metric, sel_pols, metric_label, sel_key_labels)
        win.destroy()
    
    
    tk.Button(btns, text="Plot", command=plot_now).pack(side="left", padx=8)
    tk.Button(btns, text="Cancel", command=win.destroy).pack(side="left", padx=8)


# ----------------------------
# Pol Curve Style Editor
# ----------------------------
def open_pol_style_editor():
    """Dialog to customize color/marker/linestyle per Pol. Updates POL_STYLES in place,
    so all plot types (Polarization, Power, Overlay) pick it up on the next plot."""
    marker_by_label = dict(MARKER_OPTIONS)
    marker_by_code = {v: k for k, v in MARKER_OPTIONS}
    linestyle_by_label = dict(LINESTYLE_OPTIONS)
    linestyle_by_code = {v: k for k, v in LINESTYLE_OPTIONS}

    win = tk.Toplevel(root)
    win.title("Pol Curve Styles")
    win.grab_set()

    tk.Label(win, text="Pol").grid(row=0, column=0, padx=8, pady=(8, 4))
    tk.Label(win, text="Color").grid(row=0, column=1, padx=8, pady=(8, 4))
    tk.Label(win, text="Marker").grid(row=0, column=2, padx=8, pady=(8, 4))
    tk.Label(win, text="Line style").grid(row=0, column=3, padx=8, pady=(8, 4))

    color_vars, color_btns, marker_vars, linestyle_vars = {}, {}, {}, {}

    def pick_color(pol_code):
        cur = color_vars[pol_code].get()
        _, hexval = colorchooser.askcolor(color=cur, title=f"Pick color for Pol {pol_code}")
        if hexval:
            color_vars[pol_code].set(hexval)
            color_btns[pol_code].configure(bg=hexval, activebackground=hexval)

    for idx, pol_code in enumerate(POL_CODES):
        row = 1 + idx
        style = POL_STYLES.get(pol_code, DEFAULT_POL_STYLES[pol_code])

        tk.Label(win, text=f"Pol {pol_code}").grid(row=row, column=0, padx=8, pady=3, sticky="w")

        cvar = tk.StringVar(value=style["color"])
        color_vars[pol_code] = cvar
        btn = tk.Button(win, text="   ", bg=style["color"], width=6,
                         command=lambda p=pol_code: pick_color(p))
        btn.grid(row=row, column=1, padx=8, pady=3)
        color_btns[pol_code] = btn

        mvar = tk.StringVar(value=marker_by_code.get(style["marker"], "Circle"))
        marker_vars[pol_code] = mvar
        ttk.Combobox(win, textvariable=mvar, values=[l for l, _ in MARKER_OPTIONS],
                     state="readonly", width=14).grid(row=row, column=2, padx=8, pady=3)

        lvar = tk.StringVar(value=linestyle_by_code.get(style["linestyle"], "Solid"))
        linestyle_vars[pol_code] = lvar
        ttk.Combobox(win, textvariable=lvar, values=[l for l, _ in LINESTYLE_OPTIONS],
                     state="readonly", width=12).grid(row=row, column=3, padx=8, pady=3)

    def apply_and_close():
        for pol_code in POL_CODES:
            POL_STYLES[pol_code] = {
                "color": color_vars[pol_code].get(),
                "marker": marker_by_label[marker_vars[pol_code].get()],
                "linestyle": linestyle_by_label[linestyle_vars[pol_code].get()],
            }
        set_status("Pol curve styles updated.")
        win.destroy()

    def reset_defaults():
        for pol_code in POL_CODES:
            d = DEFAULT_POL_STYLES[pol_code]
            color_vars[pol_code].set(d["color"])
            color_btns[pol_code].configure(bg=d["color"], activebackground=d["color"])
            marker_vars[pol_code].set(marker_by_code.get(d["marker"], "Circle"))
            linestyle_vars[pol_code].set(linestyle_by_code.get(d["linestyle"], "Solid"))

    btns2 = tk.Frame(win)
    btns2.grid(row=1 + len(POL_CODES), column=0, columnspan=4, pady=10)
    tk.Button(btns2, text="Apply", command=apply_and_close).pack(side="left", padx=6)
    tk.Button(btns2, text="Reset to defaults", command=reset_defaults).pack(side="left", padx=6)
    tk.Button(btns2, text="Cancel", command=win.destroy).pack(side="left", padx=6)


# ----------------------------
# README
# ----------------------------
README_FILENAME = "README.txt"
LOGO_FILENAME = "zbt-logo.png"

# LabelFrame captions: sections stand out against the ~9pt default, nested boxes
# stay one step down so the grouping is still readable.
SECTION_FONT = ("Segoe UI", 11, "bold")
SUBSECTION_FONT = ("Segoe UI", 9, "bold")

def open_readme():
    """Open README.txt (next to the program, else the bundled copy) in the default viewer."""
    path = _find_file(README_FILENAME)
    if path is None:
        messagebox.showerror("README", f"README not found:\n{os.path.join(_app_dir(), README_FILENAME)}")
        return
    try:
        os.startfile(path)
    except Exception as e:
        messagebox.showerror("README", f"Failed to open README:\n{e}")


# ----------------------------
# TK App
# ----------------------------
root = tk.Tk()
root.title("PEMFC YALLA 6")

# --- Scrollable main area -----------------------------------------------------
# Every section below lives in `content`, not in `root`, so a small window scrolls
# instead of clipping the lower buttons. Status bar and footer stay pinned.
scroll_area = tk.Frame(root)
main_canvas = tk.Canvas(scroll_area, highlightthickness=0)
main_vbar = tk.Scrollbar(scroll_area, orient="vertical", command=main_canvas.yview)
content = tk.Frame(main_canvas)
_content_id = main_canvas.create_window((0, 0), window=content, anchor="nw")
main_canvas.configure(yscrollcommand=main_vbar.set)
main_canvas.pack(side="left", fill="both", expand=True)
main_vbar.pack(side="right", fill="y")

content.bind("<Configure>", lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all")))
# Keep the inner frame as wide as the canvas, otherwise fill="x" widgets stay at
# their requested width instead of stretching with the window.
main_canvas.bind("<Configure>", lambda e: main_canvas.itemconfigure(_content_id, width=e.width))


def _on_mousewheel(event):
    # bind_all catches dialogs too; only scroll when the pointer is in the main window
    if event.widget.winfo_toplevel() is not root:
        return
    main_canvas.yview_scroll(int(-event.delta / 120), "units")


root.bind_all("<MouseWheel>", _on_mousewheel)

# ----------------------------
# Data Source
# ----------------------------
data_frame = tk.LabelFrame(content, text="Data Source", font=SECTION_FONT)
data_frame.pack(pady=8, padx=8, fill="x")

# Folder and archive side by side: two ways in without costing a second button row.
source_row = tk.Frame(data_frame)
source_row.pack(pady=(8, 4), padx=8, fill="x")

select_button = tk.Button(source_row, text="Select Folder", command=load_from_folder)
select_button.pack(side="left", fill="x", expand=True, padx=(0, 3))

archive_button = tk.Button(source_row, text="Select Archive (.rar / .zip)", command=load_from_archive)
archive_button.pack(side="left", fill="x", expand=True, padx=(3, 0))

extract_button = tk.Button(data_frame, text="Extract i@0.6 / i@0.65 + OCV + Pmax", command=extract_current)
extract_button.pack(pady=(0, 8), padx=8, fill="x")

# ----------------------------
# Plot
# ----------------------------
plot_frame = tk.LabelFrame(content, text="Plot", font=SECTION_FONT)
plot_frame.pack(pady=8, padx=8, fill="x")

tk.Label(plot_frame, text="Dataset:").pack(pady=(8, 0))
key_dropdown = ttk.Combobox(plot_frame, state="readonly")
key_dropdown.pack(pady=(0, 8), padx=8, fill="x")

# Active area + averaging (used by extraction and by every plot below)
settings_frame = tk.Frame(plot_frame)
settings_frame.pack(pady=4)
active_area_var = tk.StringVar(value="25")
tk.Label(settings_frame, text="Active area [cm²]:").grid(row=0, column=0, sticky="e", padx=(0, 6), pady=2)
tk.Entry(settings_frame, textvariable=active_area_var, width=8, justify="right").grid(row=0, column=1, sticky="w", pady=2)
block_size_var = tk.StringVar(value="10")  # default 10
tk.Label(settings_frame, text="Points per average:").grid(row=1, column=0, sticky="e", padx=(0, 6), pady=2)
tk.Entry(settings_frame, textvariable=block_size_var, width=6, justify="right").grid(row=1, column=1, sticky="w", pady=2)

# --- Polarization axis limits (leave blank for auto)
iv_axes_frame = tk.LabelFrame(plot_frame, text="Polarization curve axes", font=SUBSECTION_FONT)
iv_axes_frame.pack(pady=6, padx=8, fill="x")
iv_xmin_var = tk.StringVar(value="0")  # J min [A/cm²]
iv_xmax_var = tk.StringVar(value="4")  # J max [A/cm²]
iv_ymin_var = tk.StringVar(value="0")  # U min [V]
iv_ymax_var = tk.StringVar(value="1")  # U max [V]
tk.Label(iv_axes_frame, text="J min [A/cm²]:").grid(row=0, column=0, sticky="e", padx=6, pady=2)
tk.Entry(iv_axes_frame, textvariable=iv_xmin_var, width=8, justify="right").grid(row=0, column=1, sticky="w")
tk.Label(iv_axes_frame, text="J max [A/cm²]:").grid(row=0, column=2, sticky="e", padx=12, pady=2)
tk.Entry(iv_axes_frame, textvariable=iv_xmax_var, width=8, justify="right").grid(row=0, column=3, sticky="w")
tk.Label(iv_axes_frame, text="U min [V]:").grid(row=1, column=0, sticky="e", padx=6, pady=2)
tk.Entry(iv_axes_frame, textvariable=iv_ymin_var, width=8, justify="right").grid(row=1, column=1, sticky="w")
tk.Label(iv_axes_frame, text="U max [V]:").grid(row=1, column=2, sticky="e", padx=12, pady=2)
tk.Entry(iv_axes_frame, textvariable=iv_ymax_var, width=8, justify="right").grid(row=1, column=3, sticky="w")

# --- Power axis limits (leave blank for auto)
power_axes_frame = tk.LabelFrame(plot_frame, text="Power axes", font=SUBSECTION_FONT)
power_axes_frame.pack(pady=6, padx=8, fill="x")
power_xmin_var = tk.StringVar(value="0")  # J min [A/cm²]
power_xmax_var = tk.StringVar(value="4")  # J max [A/cm²]
power_ymin_var = tk.StringVar(value="0")  # P min [W/cm²]
power_ymax_var = tk.StringVar(value="1")  # P max [W/cm²]
tk.Label(power_axes_frame, text="J min [A/cm²]:").grid(row=0, column=0, sticky="e", padx=6, pady=2)
tk.Entry(power_axes_frame, textvariable=power_xmin_var, width=8, justify="right").grid(row=0, column=1, sticky="w")
tk.Label(power_axes_frame, text="J max [A/cm²]:").grid(row=0, column=2, sticky="e", padx=12, pady=2)
tk.Entry(power_axes_frame, textvariable=power_xmax_var, width=8, justify="right").grid(row=0, column=3, sticky="w")
tk.Label(power_axes_frame, text="P min [W/cm²]:").grid(row=1, column=0, sticky="e", padx=6, pady=2)
tk.Entry(power_axes_frame, textvariable=power_ymin_var, width=8, justify="right").grid(row=1, column=1, sticky="w")
tk.Label(power_axes_frame, text="P max [W/cm²]:").grid(row=1, column=2, sticky="e", padx=12, pady=2)
tk.Entry(power_axes_frame, textvariable=power_ymax_var, width=8, justify="right").grid(row=1, column=3, sticky="w")

# --- Toggles
show_pmax_var = tk.BooleanVar(value=True)
tk.Checkbutton(plot_frame, text="Show Pmax table (Power plot)", variable=show_pmax_var).pack(pady=(2, 8))

# --- Plot action buttons
plot_button = tk.Button(plot_frame, text="Plot Polarization", command=on_key_select)
plot_button.pack(pady=4, padx=8, fill="x")

power_plot_button = tk.Button(plot_frame, text="Plot Power (P–J)", command=on_power_plot)
power_plot_button.pack(pady=4, padx=8, fill="x")

# Marker mode for overlays: same symbol per Pol (default) vs. one symbol per dataset
vary_marker_by_dataset_var = tk.BooleanVar(value=False)
# Color mode for overlays: same color per Pol (default) vs. one color per dataset
vary_color_by_dataset_var = tk.BooleanVar(value=False)

overlay_keypol_btn = tk.Button(plot_frame, text="Overlay (Key + Pol)…", command=open_overlay_keypol_selector)
overlay_keypol_btn.pack(pady=4, padx=8, fill="x")

bar_diagram_btn = tk.Button(plot_frame, text="Bar Diagram (Extracted Values)…", command=open_bar_selector)
bar_diagram_btn.pack(pady=4, padx=8, fill="x")

pol_style_btn = tk.Button(plot_frame, text="Pol Styles…", command=open_pol_style_editor)
pol_style_btn.pack(pady=(4, 8), padx=8, fill="x")

# --- Prominent "press E" hint
hint_frame = tk.Frame(plot_frame, bg="#fff3cd", highlightbackground="#e0a800", highlightthickness=1, bd=0)
hint_frame.pack(pady=(0, 10), padx=8, fill="x")
tk.Label(
    hint_frame,
    text="Press 'E' in a plot window to edit labels",
    font=("Segoe UI", 11, "bold"),
    fg="#7a5200",
    bg="#fff3cd",
    pady=6,
).pack(fill="x")

# ----------------------------
# Results
# ----------------------------
results_frame = tk.LabelFrame(content, text="Results", font=SECTION_FONT)
results_frame.pack(pady=8, padx=8, fill="x")

export_results_button = tk.Button(results_frame, text="Export Results (CSV/XLSX)", command=export_results)
export_results_button.pack(pady=(8, 4), padx=8, fill="x")

update_excel_button = tk.Button(results_frame, text="Update Master Excel (.xlsx)", command=update_master_excel)
update_excel_button.pack(pady=(0, 4), padx=8, fill="x")

elabftw_export_button = tk.Button(results_frame, text="Export to eLabFTW (JSON)…", command=open_elabftw_export_dialog)
elabftw_export_button.pack(pady=(0, 8), padx=8, fill="x")

# ----------------------------
# Help
# ----------------------------
help_frame = tk.LabelFrame(content, text="Help", font=SECTION_FONT)
help_frame.pack(pady=8, padx=8, fill="x")

readme_button = tk.Button(help_frame, text="Open README", command=open_readme)
readme_button.pack(pady=8, padx=8, fill="x")

# --- Mini Statusbar (bottom)
status_var = tk.StringVar(value="Ready")
status_bar = tk.Label(root, textvariable=status_var, anchor="w", bd=1, relief="sunken", padx=6)
status_bar.pack(side="bottom", fill="x")

# --- Footer: discrete ZBT logo, copyright on its own line underneath so the two
# never collide when the window is narrow.
footer = tk.Frame(root)
footer.pack(side="bottom", fill="x", padx=10, pady=(4, 6))

tk.Label(
    footer,
    text="Zentrum für BrennstoffzellenTechnik GmbH 2025",
    font=("Segoe UI", 8),
    fg="#888888",
    anchor="w",
).pack(side="bottom", fill="x", pady=(2, 0))

# Keep a module-level reference: Tk does not own the image, and a local one would be
# garbage-collected immediately, leaving an empty label.
logo_image = None
try:
    _logo_path = _find_file(LOGO_FILENAME)
    if _logo_path:
        # subsample halves the logo (120x29 -> 60x14) without pulling in Pillow
        logo_image = tk.PhotoImage(file=_logo_path).subsample(2, 2)
        tk.Label(footer, image=logo_image).pack(side="left")
except Exception as e:
    print(f"[logo] Could not load {LOGO_FILENAME}: {e}", file=sys.stderr)

# Pack the scroll area last: status bar and footer are laid out first and keep
# their height, the canvas takes whatever is left.
scroll_area.pack(side="top", fill="both", expand=True)

# Open at full content size when the screen allows it, otherwise cap the height
# and let the scrollbar handle the rest.
root.update_idletasks()
_win_w = content.winfo_reqwidth() + main_vbar.winfo_reqwidth() + 4
_win_h = (content.winfo_reqheight()
          + status_bar.winfo_reqheight()
          + footer.winfo_reqheight() + 20)
root.geometry(f"{_win_w}x{min(_win_h, int(root.winfo_screenheight() * 0.9))}")
root.minsize(360, 240)

root.mainloop()
